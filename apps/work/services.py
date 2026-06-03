import os
import zipfile
import shutil
import tempfile
from contextlib import contextmanager
import re
import math
import hashlib
import fitz 
import easyocr
import cv2 
import numpy as np
from PIL import Image
from decimal import Decimal, ROUND_HALF_UP, ROUND_DOWN
from datetime import datetime
from time import perf_counter
import openpyxl
import warnings
warnings.filterwarnings("ignore", category=UserWarning) # Suppress torch/easyocr warnings
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill, Font
import requests 
from bs4 import BeautifulSoup
from django.conf import settings
import difflib # For fuzzy matching

PLATE_TOKEN_RE = re.compile(r"^\d{2,4}[A-Z]{2,4}\d{2,4}$")

# Поля с фото на странице предпросмотра (ключи — как в Excel/форме)
PREVIEW_PHOTO_FIELD_KEYS = frozenset({1, 2, 3, 4, 7, 8})


def cleanup_preview_workspace(workspace_dir):
    """Удаляет временную папку OCR (системный temp, не media)."""
    if not workspace_dir:
        return
    try:
        if os.path.isdir(workspace_dir):
            shutil.rmtree(workspace_dir, ignore_errors=True)
            print(f"[cleanup_preview_workspace] Удалено: {workspace_dir}")
    except Exception as e:
        print(f"[cleanup_preview_workspace] Ошибка: {workspace_dir}: {e}")


def _cleanup_ocr_processing_artifacts(workspace_dir):
    """После OCR удаляет zip и распаковку; кропы и preview_imgs остаются до скачивания Excel."""
    if not workspace_dir or not os.path.isdir(workspace_dir):
        return
    for name in ("upload", "extracted"):
        path = os.path.join(workspace_dir, name)
        if os.path.isdir(path):
            shutil.rmtree(path, ignore_errors=True)
    try:
        for item in os.listdir(workspace_dir):
            if item.startswith("temp_imgs_processing_obj_"):
                shutil.rmtree(os.path.join(workspace_dir, item), ignore_errors=True)
    except OSError:
        pass


def _preview_relpath(abs_path, workspace_dir):
    rel = os.path.relpath(abs_path, workspace_dir)
    return rel.replace("\\", "/")


PREVIEW_JPEG_QUALITY = 92


def _pil_rgb_to_bgr_np(pil_img):
    if pil_img.mode != "RGB":
        pil_img = pil_img.convert("RGB")
    return cv2.cvtColor(np.array(pil_img), cv2.COLOR_RGB2BGR)


def _easyocr_read_pil(pil_img, ocr_reader=None, allowlist=None):
    """OCR по PIL-изображению в памяти (без записи на диск)."""
    active = ocr_reader or reader
    if active is None:
        return []
    ocr_np = _pil_rgb_to_bgr_np(pil_img)
    try:
        if allowlist:
            return active.readtext(ocr_np, detail=1, allowlist=allowlist)
        return active.readtext(ocr_np, detail=1)
    except TypeError:
        return active.readtext(ocr_np, detail=1)


def _save_preview_jpeg(pil_img, path):
    if pil_img.mode != "RGB":
        pil_img = pil_img.convert("RGB")
    pil_img.save(path, "JPEG", quality=PREVIEW_JPEG_QUALITY, optimize=True)


def _preview_field_key(field_name: str):
    """Ключ поля (1–8) для предпросмотра по имени поля в карте координат."""
    if not field_name or "Якорь" in field_name or "Anchor" in field_name:
        return None
    if "Дата (1)" in field_name or field_name.strip().startswith("Дата"):
        return 1
    if "Марка" in field_name and "Гос" not in field_name:
        return 2
    if "Гос_номер" in field_name:
        return 3
    if "ФИО Водит" in field_name:
        return 4
    if "Кол.тон" in field_name:
        return 7
    if "Цена (8)" in field_name or field_name.strip().startswith("Цена"):
        return 8
    return None


def _is_preview_photo_field(field_name: str) -> bool:
    return _preview_field_key(field_name) is not None


def _field_key_from_image_name(img_file_lower):
    """Определяет ключ поля по транслитерированному имени JPEG (фоллбек для scan_dir)."""
    if "data_1" in img_file_lower or (
        ("date" in img_file_lower or "data" in img_file_lower) and "_1" in img_file_lower
    ):
        return 1
    if (("marka" in img_file_lower or "марка" in img_file_lower) and
            "gos" not in img_file_lower and "nomer" not in img_file_lower):
        return 2
    if ("gos" in img_file_lower and "nomer" in img_file_lower):
        return 3
    if "fio" in img_file_lower and "vodit" in img_file_lower:
        return 4
    if "kol" in img_file_lower and "ton" in img_file_lower:
        return 7
    if ("cena" in img_file_lower or "tsena" in img_file_lower or
            "_8" in img_file_lower or "(8)" in img_file_lower):
        return 8
    return None


def _merge_preview_paths_from_extract(field_images, extracted_data):
    """Добавляет пути кропов из extract_text_from_pdf (надёжнее, чем угадывание по имени файла)."""
    by_field = extracted_data.get("_preview_by_field")
    if not isinstance(by_field, dict):
        return
    for key_str, paths in by_field.items():
        try:
            field_key = int(key_str)
        except (TypeError, ValueError):
            continue
        if field_key not in PREVIEW_PHOTO_FIELD_KEYS:
            continue
        bucket = field_images.setdefault(str(field_key), [])
        for rel_path in paths:
            if rel_path and rel_path not in bucket:
                bucket.append(rel_path)


def _scan_dir_field_images(scan_dir, workspace_root, field_images, preview_paths=None):
    """Фоллбек: классифицирует JPEG в field_images (строковые ключи «1»…«8»)."""
    if not scan_dir or not os.path.isdir(scan_dir):
        return
    for img_file in os.listdir(scan_dir):
        img_path = os.path.join(scan_dir, img_file)
        if not os.path.isfile(img_path):
            continue
        lower = img_file.lower()
        if not lower.endswith((".jpg", ".jpeg", ".png")):
            continue
        rel_path = _preview_relpath(img_path, workspace_root)
        field_key = _field_key_from_image_name(lower)
        if field_key not in PREVIEW_PHOTO_FIELD_KEYS:
            continue
        if preview_paths is not None:
            preview_paths.append(rel_path)
        bucket = field_images.setdefault(str(field_key), [])
        if rel_path not in bucket:
            bucket.append(rel_path)


@contextmanager
def _ocr_temp_workspace():
    """Временная папка в системном temp; при ошибке OCR удаляется целиком."""
    path = tempfile.mkdtemp(prefix="quanta_ocr_")
    print(f"[ocr_temp_workspace] Создано: {path}")
    try:
        yield path
    except Exception:
        cleanup_preview_workspace(path)
        raise


class _TimingCollector:
    def __init__(self, enabled: bool = True):
        self.enabled = enabled
        self._next_block_id = 1
        self.blocks = []

    def start_block(self, title: str):
        if not self.enabled:
            return None
        block = {
            "block_id": self._next_block_id,
            "title": title,
            "steps": [],
            "total_seconds": None,
        }
        self._next_block_id += 1
        self.blocks.append(block)
        return block

    def set_block_total(self, block, seconds: float):
        if not block:
            return
        block["total_seconds"] = seconds

    def add_step(self, block, idx: int, description: str, seconds: float):
        if not block:
            return
        block["steps"].append(
            {
                "idx": idx,
                "description": description,
                "seconds": seconds,
            }
        )

    def print_all(self, overall_seconds=None):
        if not self.enabled:
            return
        for block in self.blocks:
            block_id = block["block_id"]
            print(f"=========Работа с файлом: {block_id}=========")
            for step in block["steps"]:
                print(f"====={step['idx']}={step['description']}")
                print(f"={step['seconds']:.3f} секунд")
        if overall_seconds is not None:
            print(f"=========ИТОГО: {overall_seconds:.3f} секунд=========")


def _format_word_groups(groups, empty_label="    (пусто)"):
    """Форматирует список (текст, высота) или ocr_debug-записей для лога."""
    if not groups:
        return empty_label
    lines = []
    for item in groups:
        if isinstance(item, (list, tuple)) and len(item) >= 2:
            text, h = item[0], item[1]
            extra = ""
            if len(item) >= 3:
                extra = f", уверенность={item[2]:.2f}"
            lines.append(f"    «{text}» — высота букв H={h}px{extra}")
        elif isinstance(item, dict):
            text = str(item.get("text", "")).strip()
            h = item.get("height", "?")
            passed = item.get("passed")
            mark = "✓" if passed else "✗"
            lines.append(f"    «{text}» — H={h}px, min={item.get('min_height', '?')} [{mark}]")
        elif item:
            lines.append(f"    «{item}»")
    return "\n".join(lines) if lines else empty_label


def _brand_cleaning_desc():
    return (
        "канал B (синий) из RGB\n"
        "бинаризация: нет\n"
        "отбор строк: 35 < высота букв H < 46 px (get_cleaned_big_3_list)\n"
        "склейка результата: через « / »"
    )


def _fio_cleaning_desc(threshold):
    min_h = CURRENT_MIN_HEIGHT_CONFIG.get("ФИО Водит. (4)", 0)
    return (
        f"канал B (синий) из RGB\n"
        f"бинаризация: порог {threshold} (пиксель > {threshold} → белый, иначе чёрный)\n"
        f"отбор ФИО: 35 < H < 500 px; min_height OCR-фильтра: {min_h}\n"
        "правила: не более 2 точек, без мусорных символов, фамилия ≥ 3 букв"
    )


def _plate_cleaning_header(threshold_options, min_height_plate):
    return (
        "канал B (синий) из RGB → серая матрица\n"
        f"список порогов бинаризации: {threshold_options}\n"
        f"пайплайны: raw, raw_x2, clahe, clahe_x2, adaptive_gauss_31_7, otsu, thr_<N>, clahe_thr_<N>\n"
        f"OCR: EasyOCR (en), allowlist A-Z0-9/\n"
        f"min_height для номера: {min_height_plate}px\n"
        "остановка: как только найдено ≥ 2 токена формата номера"
    )


def _record_field_attempt(extracted_data, field_name, entry):
    extracted_data.setdefault("_field_logs", {}).setdefault(field_name, {"attempts": []})
    attempts = extracted_data["_field_logs"][field_name]["attempts"]
    if "attempt" not in entry:
        entry["attempt"] = len(attempts) + 1
    attempts.append(entry)


def _merge_field_logs(target, source):
    if not isinstance(source, dict):
        return
    for field_name, block in source.items():
        if not isinstance(block, dict):
            continue
        target.setdefault(field_name, {"attempts": []})
        for att in block.get("attempts") or []:
            if isinstance(att, dict):
                target[field_name]["attempts"].append(dict(att))


def format_driver_ocr_details(
    file_label,
    field_logs,
    fio_meta_attempts=None,
    final_brand="",
    final_plate="",
    final_fio="",
    chosen_fio_attempt=None,
):
    """Текстовый отчёт по распознаванию полей одного водителя (для print и ошибок)."""
    lines = [
        "",
        "=" * 35,
        "Детали",
        "=" * 35,
        f"Файл: {file_label}",
        "",
    ]

    # --- Марка ---
    lines.append("===Марка АТС===")
    brand_attempts = (field_logs or {}).get("Марка", {}).get("attempts") or []
    if not brand_attempts:
        lines.append("(полный OCR не выполнялся или данные из XLSX)")
    else:
        lines.append(f"Всего попыток OCR: {len(brand_attempts)} (марка перечитывается только при полном OCR страницы)")
        for att in brand_attempts:
            lines.append(f"=попытка: {att.get('attempt', '?')}")
            if att.get("ocr_mode"):
                lines.append(f"режим: {att['ocr_mode']}")
            lines.append("Взято группа слов (все распознанные, с высотой букв):")
            lines.append(_format_word_groups(att.get("raw_groups")))
            if att.get("filtered_groups") is not None:
                lines.append("После фильтра 35<H<46:")
                lines.append(_format_word_groups([(t, "") for t in att.get("filtered_groups") or []] or None))
            lines.append("Настройка для очистки:")
            lines.append(att.get("cleaning") or _brand_cleaning_desc())
            if att.get("result"):
                lines.append(f"Итог попытки: {att['result']}")
            lines.append("")
    if final_brand:
        lines.append(f"→ Итоговая марка: {final_brand}")
    lines.append("")

    # --- Гос.номер ---
    lines.append("===Гос.номер===")
    plate_key = "Гос_номер ()"
    plate_block = (field_logs or {}).get(plate_key, {})
    plate_attempts = plate_block.get("attempts") or []
    if not plate_attempts:
        lines.append("(нет данных OCR)")
    else:
        if plate_block.get("cleaning_header"):
            lines.append("Общие настройки пайплайна:")
            lines.append(plate_block["cleaning_header"])
            lines.append("")
        chosen = plate_block.get("chosen_attempt")
        lines.append(f"Всего внутренних попыток (пайплайн×порог): {len(plate_attempts)}")
        for att in plate_attempts:
            mark = " ← ВЫБРАНО" if att.get("attempt") == chosen else ""
            if att.get("selected"):
                mark = " ← ВЫБРАНО"
            lines.append(f"=попытка: {att.get('attempt', '?')}{mark}")
            lines.append(f"пайплайн: {att.get('pipeline', '?')}")
            if att.get("threshold") is not None:
                lines.append(f"порог бинаризации: {att['threshold']}")
            lines.append(f"оценка: {att.get('score', '?')}, совпадений формата: {att.get('matches', '?')}")
            lines.append("Взято группа слов:")
            lines.append(_format_word_groups(att.get("raw_groups")))
            lines.append("")
        if plate_block.get("winning_pipeline"):
            lines.append(
                f"Выбран пайплайн: {plate_block.get('winning_pipeline')}, "
                f"порог: {plate_block.get('winning_threshold')}"
            )
        if plate_block.get("retry_failed"):
            lines.append("(ни одна попытка не дала 2 валидных части номера — взят лучший по score)")
    if final_plate:
        lines.append(f"→ Итоговый гос.номер: {final_plate}")
    lines.append("")

    # --- ФИО ---
    lines.append("===ФИО водителя===")
    fio_key = "ФИО Водит. (4)"
    fio_attempts = (field_logs or {}).get(fio_key, {}).get("attempts") or []
    fio_meta = fio_meta_attempts or []
    if not fio_attempts and not fio_meta:
        lines.append("(нет данных OCR)")
    else:
        total = max(len(fio_attempts), len(fio_meta))
        lines.append(
            f"Всего попыток с разным порогом очистки: {total} "
            f"(пороги: первая — полный OCR, далее — только поле ФИО по кэшу страницы)"
        )
        for i in range(total):
            att = fio_attempts[i] if i < len(fio_attempts) else {}
            meta = fio_meta[i] if i < len(fio_meta) else {}
            num = meta.get("attempt") or att.get("attempt") or (i + 1)
            chosen_mark = ""
            if chosen_fio_attempt is not None and num == chosen_fio_attempt:
                chosen_mark = " ← ИТОГ"
            lines.append(f"=попытка: {num}{chosen_mark}")
            if meta.get("mode"):
                lines.append(f"режим: {meta['mode']}")
            thresh = meta.get("threshold") or att.get("threshold")
            if thresh is not None:
                lines.append(f"порог очистки фото: {thresh}")
            lines.append("Взято группа слов (текст + высота букв H в px):")
            raw = att.get("raw_groups")
            if raw is None and meta.get("raw_groups"):
                raw = meta["raw_groups"]
            lines.append(_format_word_groups(raw))
            if meta.get("fio_clean") is not None:
                lines.append(f"После clean_fio_raw: «{meta.get('fio_clean', '')}»")
            if meta:
                rules = "да" if meta.get("rules_passed") else "нет"
                lines.append(f"правила ФИО: {rules}")
                if meta.get("rules_fail_reason"):
                    lines.append(f"причина отказа: {meta['rules_fail_reason']}")
                if meta.get("t2_match") or meta.get("t3_match"):
                    lines.append("совпадение с файлом ЭСФ/СНТ: да")
                elif meta.get("rules_passed"):
                    lines.append("совпадение с файлом ЭСФ/СНТ: нет (продолжаем перебор)")
            lines.append("Настройки для очистки:")
            lines.append(att.get("cleaning") or _fio_cleaning_desc(thresh or "?"))
            lines.append("")
    if final_fio:
        lines.append(f"→ Итоговое ФИО: {final_fio}")
    lines.append("=" * 35)
    return "\n".join(lines)


class NetworkError(Exception):
    def __init__(self, user_message, technical_details):
        self.user_message = user_message
        self.technical_details = technical_details
        super().__init__(user_message)

NBKR_URL = "https://www.nbkr.kg/index1.jsp?item=1562&lang=RUS&valuta_id=15"

def selected_usa_dollar(soup):
    option = soup.find("option", value="15")
    # print(f"[get_current_dollar_rate.selected_usa_dollar] option found: {option}")
    value_text = '<option selected="" value="15">1 Доллар США'
    if value_text in str(option):
        return True
    else:
        return False

def get_curs(soup, date):
    table = soup.find_all("tr")

    for tr in table:
        if f"{date}" in str(tr):
            td = tr.find("td", class_="stat-right")
            value = td.get_text()
            rate_decimal = Decimal(value.replace(",", "."))
            rate_truncated = (rate_decimal * Decimal("100")).quantize(Decimal("1"), rounding=ROUND_DOWN) / Decimal("100")
            print("Курс:", rate_truncated)
            return rate_truncated

def get_current_dollar_rate(date_str=None):
    try:
        resp = requests.get(NBKR_URL, timeout=10)
        html = resp.text
        # print(f"[get_current_dollar_rate] Fetched NBKR_URL status={resp.status_code} length={len(html)} for date={date_str}")
    except requests.RequestException as e:
        # print(f"[get_current_dollar_rate] RequestException: {e} (url={NBKR_URL})")
        user_message = "Проверьте подключение к интернету"
        technical_details = f"Ошибка при подключении к сайту НБКР: {str(e)}"
        raise NetworkError(user_message, technical_details)

    soup = BeautifulSoup(html, "html.parser")
    
    if not selected_usa_dollar(soup):
        raise Exception("На странице НБКР не выбран доллар США")
    
    if not date_str:
        print("[get_current_dollar_rate] date_str is empty or None")
        raise Exception("Дата не указана")
    
    rate = get_curs(soup, date_str)
    
    if rate:
        print(f"Получен курс доллара США на {date_str}: {rate}")
        return rate
    else:
        print(f"[get_current_dollar_rate] rate not found for date {date_str}")
        raise Exception(f"Не удалось найти курс на дату {date_str}")
# Настройки авто-замены для Марки АТС
# if_have (если в тексте есть это слово -> заменить всю строку на указанное значение)
BRAND_CONTAIN_MAP = {
    "scania": "Scania",
    "volvo": "Volvo",
}

# if_it (точечная замена подстроки для исправления опечаток в буквах)
BRAND_SUBSTRING_MAP = {
    r'(?i)(?<!v)olvo': 'Volvo',
    r'(?i).*olvo.*': 'Volvo',
    r'(?i).*scania.*': 'Scania',
}

def clean_brand_name(brand_text):
    if not brand_text:
        return brand_text
    
    # 1. Точечная замена подстрок (if_it)
    for pattern, replacement in BRAND_SUBSTRING_MAP.items():
        brand_text = re.sub(pattern, replacement, brand_text)
        
    # 2. Поиск ключевых слов для замены всей строки (if_have)
    brand_text_lower = brand_text.lower()
    for key, replacement in BRAND_CONTAIN_MAP.items():
        if key in brand_text_lower:
            return replacement
            
    return brand_text.strip()

# 2026
FIELDS_MAP_TYPE_1 = {
    "ФИО Водит. (4)": (850, 2305, 500, 150),
    "Кол.тон (7)": (1470, 1275, 300, 200),
    "Марка": (350, 2575, 350, 120),
    "Гос_номер ()": (-80, 2580, 600, 200),
    "Якорь (1)": (0, 0, 400, 500),
}

FIELDS_MAP_TYPE_2 = {
    "Цена (8)": (1585, 2160, 250, 160),
    "№ счет факт (Инвойс) (16)": (485, 190, 500, 60),
}

FIELDS_MAP_TYPE_3 = {
    "№ сопров.накл. KZ (15)": (360, 250, 330, 200),
    "Дата сопр.накл (13)": (360, 250, 330, 200)
}

FIELDS_MAP_TYPE_2_PAGE_2 = {
    "Цена (8) Alt": (1500, 100, 150, 120),
}

MIN_HEIGHT_CONFIG = {
    "ФИО Водит. (4)": 20,
    "Марка_Гос_номер ()": 38,
}
CURRENT_MIN_HEIGHT_CONFIG = MIN_HEIGHT_CONFIG

LEGACY_FIELDS_MAP_TYPE_1 = {
    "ФИО Водит. (4)": (850, 2450, 500, 150),
    "Кол.тон (7)": (1500, 1390, 300, 80),
    "Марка": (380, 2730, 350, 70),
    "Гос_номер ()": (-80, 2730, 700, 170),
    "Якорь (1)": (0, 0, 500, 1000),
}

LEGACY_FIELDS_MAP_TYPE_2 = {
    "Цена (8)": (1450, 2150, 250, 160),
    "№ счет факт (Инвойс) (16)": (500, 200, 500, 60),
}

LEGACY_FIELDS_MAP_TYPE_3 = {
    "№ сопров.накл. KZ (15)": (360, 250, 330, 200),
    "Дата сопр.накл (13)": (360, 250, 330, 200)
}

LEGACY_FIELDS_MAP_TYPE_2_PAGE_2 = {
    "Цена (8) Alt": (1500, 100, 150, 120),
}

LEGACY_MIN_HEIGHT_CONFIG = {
    "ФИО Водит. (4)": 20,
    "Марка_Гос_номер ()": 38,
}

def get_maps_by_zip_name(zip_filename):
    """
    Выбирает координаты по году в имени архива.
    Формат имени: dd-mm-yyyy_hh-mm-ss (например, 04-04-2026_13-48-20.zip).
    Для года > 2025 используются текущие карты, иначе legacy.
    """
    match = re.search(r'(\d{2})-(\d{2})-(\d{4})_\d{2}-\d{2}-\d{2}', zip_filename or "")
    year = int(match.group(3)) if match else 2026

    use_legacy = year <= 2025
    print(f"[MAP SELECT] ZIP: {zip_filename}, year={year}, use_legacy={use_legacy}")

    if use_legacy:
        return {
            "type_1": LEGACY_FIELDS_MAP_TYPE_1,
            "type_2": LEGACY_FIELDS_MAP_TYPE_2,
            "type_3": LEGACY_FIELDS_MAP_TYPE_3,
            "type_2_page_2": LEGACY_FIELDS_MAP_TYPE_2_PAGE_2,
            "min_height": LEGACY_MIN_HEIGHT_CONFIG,
        }

    return {
        "type_1": FIELDS_MAP_TYPE_1,
        "type_2": FIELDS_MAP_TYPE_2,
        "type_3": FIELDS_MAP_TYPE_3,
        "type_2_page_2": FIELDS_MAP_TYPE_2_PAGE_2,
        "min_height": MIN_HEIGHT_CONFIG,
    }

try:
    # Main OCR reader (ru/en + optional kk for Kazakh)
    try:
        reader = easyocr.Reader(["ru", "en", "kk"], gpu=False)
    except Exception as e_kk:
        print(f"EasyOCR init without kk (fallback ru/en). reason={e_kk}")
        reader = easyocr.Reader(["ru", "en"], gpu=False)
except Exception as e:
    print(f"Error initializing EasyOCR: {e}")
    reader = None

try:
    # Plate-specific reader: English only (less confusion with Cyrillic)
    plate_reader = easyocr.Reader(["en"], gpu=False)
except Exception as e:
    print(f"Error initializing plate EasyOCR reader: {e}")
    plate_reader = None

def deskew_image(img_cv):
    try:
        # Оптимизация скорости:
        # Canny/Hough считаются по полному изображению и на листах dpi=300 это очень тяжело.
        # Детектируем угол на уменьшенной копии, а потом поворачиваем исходное изображение.
        h, w = img_cv.shape[:2]
        scale = 1.0
        max_dim = max(h, w)
        if max_dim > 2000:
            scale = 0.5
        if scale != 1.0:
            img_small = cv2.resize(
                img_cv,
                (int(w * scale), int(h * scale)),
                interpolation=cv2.INTER_AREA,
            )
        else:
            img_small = img_cv

        gray = cv2.cvtColor(img_small, cv2.COLOR_BGR2GRAY)
        edges = cv2.Canny(gray, 50, 150, apertureSize=3)

        min_line_len = max(10, int(100 * scale))
        lines = cv2.HoughLinesP(
            edges,
            1,
            np.pi / 180,
            threshold=100,
            minLineLength=min_line_len,
            maxLineGap=10,
        )

        angles = []
        if lines is not None:
            # Ограничиваем количество линий для ускорения
            for i, line in enumerate(lines):
                if i >= 200:
                    break
                x1, y1, x2, y2 = line[0]

                angle_rad = math.atan2(y2 - y1, x2 - x1)
                angle_deg = math.degrees(angle_rad)

                if abs(angle_deg) > 45:
                    if angle_deg > 0:
                        deviation = angle_deg - 90
                    else:
                        deviation = angle_deg + 90

                    angles.append(deviation)

        if not angles:
            return img_cv

        median_angle = np.median(angles)

        if abs(median_angle) < 0.1:
            return img_cv

        print(f" [Deskew] Обнаружен перекос: {median_angle:.2f} градусов. Исправляем...")

        (h, w) = img_cv.shape[:2]
        center = (w // 2, h // 2)

        M = cv2.getRotationMatrix2D(center, median_angle, 1.0)

        rotated = cv2.warpAffine(
            img_cv, M, (w, h),
            flags=cv2.INTER_CUBIC,
            borderMode=cv2.BORDER_CONSTANT,
            borderValue=(255, 255, 255)
        )

        return rotated
    except Exception as e:
        print(f" [Deskew Error] Не удалось выровнять: {e}")
        return img_cv

class DataCleaner:
    FIO_DISALLOWED_CHARS = set('()=-+!"№;%:?*/\\":?><,\'0123456789')

    @staticmethod
    def replace_ruble(text):
        if not text:
            return text
        return text.replace('₽', 'Р')

    @staticmethod
    def clean_1(text, context):
        if text:
            match = re.search(r'\b(\d{2}\.\d{2}\.\d{4})\b', text)
            if match:
                return match.group(1)

        surname = context.get('surname')
        if surname:
            all_files = context.get('type_2_files', []) + context.get('type_3_files', [])
            for fpath in all_files:
                fname = os.path.basename(fpath)
                if surname.lower() in fname.lower():
                    match_file = re.search(r'(\d{2}\.\d{2}\.\d{4})', fname)
                    if match_file:
                        return match_file.group(1)

        zip_name = context.get('zip_filename', '')
        match_zip = re.search(r'(\d{2}-\d{2}-\d{4})', zip_name)
        if match_zip:
            return match_zip.group(1).replace('-', '.')

        return text

    @staticmethod
    def get_cleaned_big_3_list(data):
        if not isinstance(data, list):
            return []

        filtered = []
        for text, h in data:
            if 35 < h < 55:
                filtered.append(text.strip())

        while filtered and filtered[0] in ["25", "26", "27", "28", "29", "30"]:
            filtered.pop(0)

        return filtered

    @staticmethod
    def clean_plate_text(text):
        t = text.strip().upper()
        
        cyr_to_lat = {
            'А': 'A', 'В': 'B', 'Е': 'E', 'К': 'K', 'М': 'M', 'Н': 'H',
            'О': 'O', 'Р': 'P', 'С': 'C', 'Т': 'T', 'У': 'Y', 'Х': 'X'
        }
        for cyr, lat in cyr_to_lat.items():
            t = t.replace(cyr, lat)

        t = t.replace('L', 'I') 
        t = t.replace('|', 'I')
        t = t.replace('I', 'I') 
        
        t = t.replace('O', '0')
        t = t.replace('S', '5')
        
        t = re.sub(r'[^A-Z0-9/]', '', t)
        
        return t

    @staticmethod
    def is_plate_token_like(token: str) -> bool:
        if not token:
            return False
        cleaned = DataCleaner.clean_plate_text(token)
        return bool(PLATE_TOKEN_RE.match(cleaned))

    @staticmethod
    def is_plate_result_like(raw_items) -> bool:
        """
        raw_items: list[tuple[text, height]] (already height-filtered for plate field)
        Returns True if at least one token looks like a plate.
        """
        if not isinstance(raw_items, list):
            return False
        for text, _h in raw_items:
            if DataCleaner.is_plate_token_like(str(text)):
                return True
        return False

    @staticmethod
    def clean_2(data, context): return ""
    @staticmethod
    def clean_3(data, context): return ""
    
    @staticmethod
    def clean_fio_raw(data, context):
        if not isinstance(data, list):
            return DataCleaner.replace_ruble(str(data).strip()), DataCleaner.replace_ruble(str(data).strip())

        filtered_items = []
        for text, h in data:
            if re.search(r'\d{2}\.\d{2}\.\d{4}', text):
                continue
                
            if 35 < h < 500:
                filtered_items.append((text.strip(), h))
        
        while filtered_items and filtered_items[0][0] in ["21", "22", "23", "24"]:
            filtered_items.pop(0)
        
        if filtered_items:
            first_item = filtered_items[0]
            t, h = first_item
            
            if t.endswith(':'):
                t = t[:-1] + '.'
            if t.endswith('-'):
                t = t[:-1] + '.'
            
            t = DataCleaner.replace_ruble(t)
            
            if not t.endswith('.'):
                t += '.'
            
            return t, t
            
        return "", ""

    @staticmethod
    def fio_rules_ok(fio_text: str) -> bool:
        """
        Проверка распознанного ФИО по правилам:
        - не более 2 точек (например, И.И. в середине + финальная точка не должны "разрастаться")
        - не должны встречаться "мусорные" символы из заданного списка
        """
        if fio_text is None:
            return False
        t = str(fio_text)
        if not t.strip():
            return False

        if t.count(".") > 2:
            return False

        if any(ch in t for ch in DataCleaner.FIO_DISALLOWED_CHARS):
            return False

        # Часто OCR "подхватывает" текст со штампа/печати.
        # В таких случаях первая "фамилия" бывает очень короткой (например, "ТО").
        # Это мешает матчить по файлам, поэтому отбрасываем слишком короткое первое слово.
        words = [w for w in re.split(r"\s+", t.strip()) if w]
        if words:
            first = words[0]
            # Оставляем только буквы/точки как в sanitize_surname
            first_clean = re.sub(r"[^A-Za-zА-Яа-яЁёӘәҒғҚқҢңӨөҰұҮүІі\.]", "", first)
            first_clean = first_clean.replace(".", "")
            if len(first_clean) < 3:
                return False

        return True

    @staticmethod
    def clean_4(text, context): return text.strip("'") if text else text
    @staticmethod
    def clean_5(text, context): return str(text).strip("'") if text else text
    @staticmethod
    def clean_6(text, context): return text
    
    @staticmethod
    def clean_7(text, context): 
        match = re.search(r'(\d{2}\s?\d{3})\s*нетто', text, re.IGNORECASE)
        if match:
            return match.group(1).replace(' ', '')
        return text

    @staticmethod
    def clean_8(text, context): 
        cleaned_text = DataCleaner.replace_ruble(str(text)).strip() if text is not None else text
        if not cleaned_text:
            return cleaned_text

        # OCR sometimes reads prices as 5 digits with extra "10" prefix (e.g. 10510 -> 510).
        digits_only = re.sub(r"\D", "", cleaned_text)
        if len(digits_only) == 5 and digits_only.startswith("10"):
            return digits_only[2:]

        return cleaned_text
    @staticmethod
    def clean_9(text, context): return text
    @staticmethod
    def clean_10(text, context): return text
    @staticmethod
    def clean_11(text, context): return text
    @staticmethod
    def clean_12(text, context): return text
    @staticmethod
    def clean_13(text, context): return str(text).strip("'") if text else text
    
    @staticmethod
    def clean_14(text, context): 
        filename = context.get('filename', '')
        # Убираем расширение
        base_name = os.path.splitext(filename)[0]
        # Убираем точки (мусор)
        base_name = base_name.replace('.', '')
        # Убираем суффиксы CMP, СМП, СМР и т.д.
        base_name = re.sub(r'\s*(cmp|смп|смр|cmr)', '', base_name, flags=re.IGNORECASE)
        
        return base_name.strip("'").strip()
    
    @staticmethod
    def clean_15(text, context):
        match = re.search(r'(KZ-SNT-[\w-]+(?:\s+[\w-]+)*)', text)
        if match:
            return match.group(1).replace(" ", "")
        return text

    @staticmethod
    def clean_16(text, context): 
        return DataCleaner.replace_ruble(text)

    @staticmethod
    def clean_marka_gos_number(data, context):
        cleaned_list = DataCleaner.get_cleaned_big_3_list(data)
        return " / ".join(cleaned_list)

def normalize_surname(surname):
    if not surname:
        return []
    variants = {surname}
    replacements = {
        'i': 'и', 'o': 'о', 'a': 'а', 'e': 'е', 'c': 'с', 'p': 'р', 
        'y': 'у', 'x': 'х', 'H': 'Н', 'K': 'К', 'M': 'М', 'B': 'В', 'T': 'Т'
    }
    new_surname = surname
    for lat, cyr in replacements.items():
        new_surname = new_surname.replace(lat, cyr)
    variants.add(new_surname)

    # Добавляем латинскую версию фамилии для поиска по файлам (для казахских/русских имен)
    latin_variant = []
    for char in surname:
        if char in CYRILLIC_TO_LATIN:
            latin_variant.append(CYRILLIC_TO_LATIN[char])
        else:
            latin_variant.append(char)
    variants.add("".join(latin_variant))

    return list(variants)

def sanitize_surname(raw_surname):
    """
    Оставляет в фамилии только буквы (рус/каз/англ) и точку.
    Убирает мусорные символы: , ] [ % и т.п.
    """
    if not raw_surname:
        return ""
    cleaned = re.sub(r"[^A-Za-zА-Яа-яЁёӘәҒғҚқҢңӨөҰұҮүІі\.]", "", str(raw_surname))
    cleaned = re.sub(r"\.{2,}", ".", cleaned)
    return cleaned.strip(". ")

def safe_decimal(value, field_name):
    if not value:
        return Decimal("0")
    cleaned = ""
    for ch in value:
        if ch.isdigit() or ch in ".,":
            cleaned += ch
    cleaned = cleaned.replace(",", ".")
    if cleaned == "":
        return Decimal("0")
    try:
        return Decimal(cleaned)
    except Exception as e:
        print(f"Ошибка Decimal для '{field_name}': '{value}' в '{cleaned}': {e}")
        return Decimal("0")

CYRILLIC_TO_LATIN = {
    'А': 'A', 'Б': 'B', 'В': 'V', 'Г': 'G', 'Д': 'D', 'Е': 'E', 'Ё': 'E',
    'Ж': 'Zh', 'З': 'Z', 'И': 'I', 'Й': 'Y', 'К': 'K', 'Л': 'L', 'М': 'M',
    'Н': 'N', 'О': 'O', 'П': 'P', 'Р': 'R', 'С': 'S', 'Т': 'T', 'У': 'U',
    'Ф': 'F', 'Х': 'H', 'Ц': 'Ts', 'Ч': 'Ch', 'Ш': 'Sh', 'Щ': 'Sch',
    'Ъ': '', 'Ы': 'Y', 'Ь': '', 'Э': 'E', 'Ю': 'Yu', 'Я': 'Ya',
    'Ә': 'A', 'Ғ': 'G', 'Қ': 'Q', 'Ң': 'N', 'Ө': 'O', 'Ұ': 'U', 'Ү': 'U', 'H': 'H', 'І': 'I',
    'а': 'a', 'б': 'b', 'в': 'v', 'г': 'g', 'д': 'd', 'е': 'e', 'ё': 'e',
    'ж': 'zh', 'з': 'z', 'и': 'i', 'й': 'y', 'к': 'k', 'л': 'l', 'м': 'm',
    'н': 'n', 'о': 'o', 'п': 'p', 'р': 'r', 'с': 's', 'т': 't', 'у': 'u',
    'ф': 'f', 'х': 'h', 'ц': 'ts', 'ч': 'ch', 'ш': 'sh', 'щ': 'sch',
    'ъ': '', 'ы': 'y', 'ь': '', 'э': 'e', 'ю': 'yu', 'я': 'ya',
    'ә': 'a', 'ғ': 'g', 'қ': 'q', 'ң': 'n', 'ө': 'o', 'ұ': 'u', 'ү': 'u', 'h': 'h', 'і': 'i'
}

def get_safe_filename(original_name, field_name):
    """
    Создает безопасное ASCII-имя файла на основе MD5-хеша исходного имени.
    Это решает проблему с кириллицей в именах файлов в Windows.
    """
    
    def transliterate(text):
        result = []
        for char in text:
            if char in CYRILLIC_TO_LATIN:
                result.append(CYRILLIC_TO_LATIN[char])
            elif char.isalnum() or char in '_-':
                result.append(char)
            else:
                result.append('_')
        return ''.join(result)
    
    combined = f"{original_name}_{field_name}"
    hash_obj = hashlib.md5(combined.encode('utf-8'))
    hash_hex = hash_obj.hexdigest()[:8]
    safe_base = os.path.splitext(os.path.basename(original_name))[0]
    safe_base = transliterate(safe_base)
    safe_base = re.sub(r'[^a-zA-Z0-9_-]', '_', safe_base)
    if len(safe_base) > 20:
        safe_base = safe_base[:20]
    
    safe_field = transliterate(field_name)
    safe_field = re.sub(r'[^a-zA-Z0-9_-]', '_', safe_field)
    safe_field = re.sub(r'_+', '_', safe_field).strip('_')
    
    return f"{safe_base}_{hash_hex}_{safe_field}.jpg"

def extract_name_from_filename(filename):
    base = os.path.splitext(os.path.basename(filename))[0]
    base = re.sub(r'^(эсф|снт|электронный\s*счет(-)?\s*фактура|счет-фактура|сопроводительная\s*накладная\s*(на)?\s*товары)\s*', '', base, flags=re.IGNORECASE)
    base = base.replace('_', ' ').strip()
    return base

def extract_surname_from_filename(filename):
    base = os.path.splitext(os.path.basename(filename))[0]
    base = base.replace('_', ' ').strip()
    words = [w for w in re.split(r'\s+', base) if w]
    if words:
        last_word = words[-1]
        cleaned = re.sub(r'[^A-Za-zА-Яа-яЁёӘәҒғҚқҢңӨөҰұҮүІі-]', '', last_word)
        if len(cleaned) >= 3:
            return cleaned
    return None

def extract_text_from_pdf(
    pdf_path,
    coords_map,
    save_dir,
    apply_deskew=False,
    page_num=0,
    fio_threshold=170,
    page_cache_holder=None,
    page_cache_image=None,
    anchor_offset_cached=None,
    timing_collector=None,
    extract_mode_label=None,
    save_preview_files=True,
    preview_workspace_root=None,
):
    """
    Извлекает поля по координатам с первой страницы PDF (или с кэшированного изображения страницы).

    Для повторных попыток ФИО: page_cache_image или page_cache_holder['image'] (RAM, без PNG на диск).
    OCR полей — из памяти; JPEG на диск только для предпросмотра (save_preview_files).
    """
    extracted_data = {}
    extracted_data["ocr_debug"] = {}
    extracted_data["_field_logs"] = {}
    doc = None
    timing_block = timing_collector.start_block(os.path.basename(pdf_path)) if timing_collector else None
    timing_total_start = perf_counter()
    step_idx = 1
    try:
        os.makedirs(save_dir, exist_ok=True)

        cached_pil = page_cache_image
        if cached_pil is None and isinstance(page_cache_holder, dict):
            cached_pil = page_cache_holder.get("image")
        used_page_cache = cached_pil is not None
        if used_page_cache:
            t0 = perf_counter()
            img_full = cached_pil.copy()
            t1 = perf_counter()
            if timing_collector and timing_block:
                timing_collector.add_step(
                    timing_block,
                    step_idx,
                    "страница из RAM-кэша (без диска)",
                    t1 - t0,
                )
                step_idx += 1
        else:
            # 1) PDF -> pixmap (самый дорогой кусок по времени/памяти)
            t0 = perf_counter()
            doc = fitz.open(pdf_path)
            if page_num >= len(doc):
                print(f"[extract_text_from_pdf] Page {page_num} does not exist in {pdf_path}")
                doc.close()
                doc = None
                return {}
            page = doc.load_page(page_num)

            # Оптимизация рендера:
            # - alpha=False: меньше каналов/памяти
            # - colorspace=RGB: предсказуемый формат
            try:
                pix = page.get_pixmap(dpi=300, alpha=False, colorspace=fitz.csRGB)
            except TypeError:
                # Фоллбек для старых версий PyMuPDF
                pix = page.get_pixmap(dpi=300)
            t1 = perf_counter()
            if timing_collector and timing_block:
                timing_collector.add_step(
                    timing_block,
                    step_idx,
                    "PDF → PNG (рендер растра dpi=300, alpha=False)",
                    t1 - t0,
                )
                step_idx += 1

            # 2) pixmap -> OpenCV BGR
            t0 = perf_counter()
            img_np = np.frombuffer(pix.samples, dtype=np.uint8).reshape(pix.height, pix.width, pix.n)
            if pix.n == 3:
                img_cv = cv2.cvtColor(img_np, cv2.COLOR_RGB2BGR)
            elif pix.n == 4:
                img_cv = cv2.cvtColor(img_np, cv2.COLOR_RGBA2BGR)
            else:
                img_cv = cv2.cvtColor(img_np, cv2.COLOR_GRAY2BGR)
            t1 = perf_counter()
            if timing_collector and timing_block:
                timing_collector.add_step(
                    timing_block,
                    step_idx,
                    "конвертация pixmap → OpenCV (BGR)",
                    t1 - t0,
                )
                step_idx += 1

            # 3) deskew (дорого на больших листах; уже оптимизировали внутри deskew_image)
            if apply_deskew:
                t0 = perf_counter()
                img_cv = deskew_image(img_cv)
                t1 = perf_counter()
                if timing_collector and timing_block:
                    timing_collector.add_step(
                        timing_block,
                        step_idx,
                        "deskew (поиск угла + поворот)",
                        t1 - t0,
                    )
                    step_idx += 1

            # 4) OpenCV -> PIL
            t0 = perf_counter()
            img_rgb = cv2.cvtColor(img_cv, cv2.COLOR_BGR2RGB)
            img_full = Image.fromarray(img_rgb)
            t1 = perf_counter()
            if timing_collector and timing_block:
                timing_collector.add_step(
                    timing_block,
                    step_idx,
                    "конвертация OpenCV → PIL (RGB)",
                    t1 - t0,
                )
                step_idx += 1

            if isinstance(page_cache_holder, dict):
                page_cache_holder["image"] = img_full.copy()

        offset_x = 0
        offset_y = 0

        # Check for explicit Anchor definition in the map
        anchor_rect = None
        anchor_key = None
        for k, v in coords_map.items():
            if "Якорь" in k or "Anchor" in k:
                anchor_rect = v
                anchor_key = k
                break

        anchor_t0 = perf_counter()
        use_cached_anchor = anchor_offset_cached is not None
        if use_cached_anchor:
            try:
                offset_x = int(anchor_offset_cached[0])
                offset_y = int(anchor_offset_cached[1])
            except (TypeError, IndexError, ValueError):
                offset_x, offset_y = 0, 0
        elif anchor_rect:
            # ONLY run anchor logic if the map has an anchor key
            ax, ay, aw, ah = anchor_rect
            # Safety checks
            ax = max(0, ax)
            ay = max(0, ay)
            anchor_crop = img_full.crop((ax, ay, ax + aw, ay + ah))

            if reader is not None:
                try:
                    print("=" * 35)
                    print(f"[ANCHOR] OCR области якоря «БИН» ({os.path.basename(pdf_path)})")
                    print("=" * 35)

                    # Оптимизация Anchor OCR:
                    # уменьшаем картинку (если она большая), чтобы EasyOCR тратил меньше времени.
                    # Важно: bbox из EasyOCR будет в координатах уменьшенного изображения, поэтому пересчитываем обратно.
                    anchor_for_ocr = anchor_crop
                    max_dim = max(anchor_for_ocr.size[0], anchor_for_ocr.size[1])
                    scale = 1.0
                    if max_dim > 900:
                        scale = 900.0 / float(max_dim)
                        new_w = max(1, int(anchor_for_ocr.size[0] * scale))
                        new_h = max(1, int(anchor_for_ocr.size[1] * scale))
                        anchor_for_ocr = anchor_for_ocr.resize((new_w, new_h), Image.BILINEAR)

                    # Fix: Pass numpy array to EasyOCR to avoid OpenCV 'can't open/read file' error with Cyrillic paths
                    anchor_np = cv2.cvtColor(np.array(anchor_for_ocr), cv2.COLOR_RGB2BGR)
                    anchor_results = reader.readtext(anchor_np, detail=1)

                    print("сырые данные из этого фото которые были взяты")
                    print("=" * 35)

                    if not anchor_results:
                        print("    (Пусто, OCR ничего не увидел)")
                    else:
                        for (bbox, text, prob) in anchor_results:
                            h = int(bbox[2][1] - bbox[0][1])
                            print(f"    - '{text}' (H: {h}, Prob: {prob:.2f})")

                    target_anchor_text = "БИН"  # якорь

                    found_anchor = False
                    for (bbox, text, prob) in anchor_results:
                        if target_anchor_text in text:
                            # bbox координаты относятся к anchor_for_ocr (возможно уменьшено)
                            local_x = float(bbox[0][0])
                            local_y = float(bbox[0][1])
                            if scale != 1.0:
                                local_x = local_x / scale
                                local_y = local_y / scale

                            offset_x = local_x
                            offset_y = local_y

                            print(f"[ANCHOR DEBUG] Якорь мы искали '{target_anchor_text}' нашли: '{text}'")
                            print(f"[ANCHOR DEBUG] Координаты которые мы ожидаем: х=0, у=0")
                            print(f"[ANCHOR DEBUG] Координаты найденного якоря: x={offset_x}, y={offset_y}")
                            print(f"[ANCHOR DEBUG] Расчет смещения: сдвиг по х={offset_x}, сдвиг по у={offset_y}")

                            found_anchor = True
                            break

                    if not found_anchor:
                        print(f"[ANCHOR DEBUG] Якорь '{target_anchor_text}' не найден в {anchor_rect}. Смещение (0,0).")
                except Exception as e:
                    print(f"[ANCHOR] Error: {e}")
        anchor_t1 = perf_counter()
        if timing_collector and timing_block:
            if use_cached_anchor:
                anchor_desc = "якорь из кэша (offset_x/offset_y)"
            elif anchor_rect:
                anchor_desc = "найти якорь (OCR EasyOCR по области БИН)"
            else:
                anchor_desc = "найти якорь (карта без якоря)"
            timing_collector.add_step(timing_block, step_idx, anchor_desc, anchor_t1 - anchor_t0)
            step_idx += 1

        extracted_data["_anchor_offset"] = (int(offset_x), int(offset_y))

        quiet_anchor_apply = used_page_cache

        for field_name, (x0, y0, w, h) in coords_map.items():
            if field_name == anchor_key:
                continue

            field_t0 = perf_counter()
            field_img_filename = None
                
            if offset_x != 0 or offset_y != 0:
                 x0 = max(0, x0 + offset_x)
                 y0 = max(0, y0 + offset_y)
                 if anchor_key and not quiet_anchor_apply:
                     print(f"[ANCHOR DEBUG] Применяем к полю '{field_name}'... New coords: ({x0}, {y0})")

            x1 = min(x0 + w, img_full.width)
            y1 = min(y0 + h, img_full.height)
            
            crop_img = img_full.crop((x0, y0, x1, y1))
            plate_ocr_precomputed = False
            selected_threshold = None
            selected_plate_pipeline = None
            allowlist_plate = "ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789/"
            
            if any(x in field_name for x in ["ФИО Водит.", "Марка", "Гос_номер"]):
                r, g, b = crop_img.split()
                crop_img = b 
                
                if "ФИО Водит." in field_name:
                    threshold = fio_threshold
                    crop_img = crop_img.point(lambda p: 255 if p > threshold else 0)
                elif "Гос_номер" in field_name:
                    threshold_options = [165, 162, 160, 157, 155, 167, 170, 172, 174]
                    best_results = []
                    successful_results = []
                    selected_threshold = threshold_options[0]
                    min_height_plate = CURRENT_MIN_HEIGHT_CONFIG.get(field_name, 0)

                    def _ocr_read_plate(img_np):
                        active = plate_reader or reader
                        if active is None:
                            return []
                        try:
                            return active.readtext(img_np, detail=1, allowlist=allowlist_plate)
                        except TypeError:
                            return active.readtext(img_np, detail=1)

                    def _score_plate_results(candidate_results):
                        if not candidate_results:
                            return 0.0, 0
                        filtered = []
                        probs = []
                        for (bbox, text, prob) in candidate_results:
                            height = int(((bbox[3][1] - bbox[0][1]) + (bbox[2][1] - bbox[1][1])) / 2)
                            if height >= min_height_plate:
                                filtered.append(str(text))
                                try:
                                    probs.append(float(prob))
                                except Exception:
                                    pass

                        matches = 0
                        for t in filtered:
                            cleaned = DataCleaner.clean_plate_text(t)
                            parts = [p for p in cleaned.split("/") if p]
                            if parts:
                                for p in parts:
                                    if DataCleaner.is_plate_token_like(p):
                                        matches += 1
                            else:
                                if DataCleaner.is_plate_token_like(cleaned):
                                    matches += 1

                        avg_prob = (sum(probs) / len(probs)) if probs else 0.0
                        bonus_two = 5.0 if matches >= 2 else 0.0
                        return matches * 10.0 + bonus_two + avg_prob, matches

                    def _plate_preprocess_variants(gray_np):
                        variants = []
                        g = gray_np.astype(np.uint8)
                        variants.append(("raw", None, g))

                        try:
                            g2 = cv2.resize(g, None, fx=2.0, fy=2.0, interpolation=cv2.INTER_CUBIC)
                            variants.append(("raw_x2", None, g2))
                        except Exception:
                            pass

                        try:
                            clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
                            g_clahe = clahe.apply(g)
                            variants.append(("clahe", None, g_clahe))
                            g_clahe2 = cv2.resize(g_clahe, None, fx=2.0, fy=2.0, interpolation=cv2.INTER_CUBIC)
                            variants.append(("clahe_x2", None, g_clahe2))
                        except Exception:
                            pass

                        try:
                            adapt = cv2.adaptiveThreshold(
                                g, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 31, 7
                            )
                            variants.append(("adaptive_gauss_31_7", None, adapt))
                        except Exception:
                            pass

                        try:
                            _t, otsu = cv2.threshold(g, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
                            variants.append(("otsu", None, otsu))
                        except Exception:
                            pass

                        for thr in threshold_options:
                            try:
                                _t, bin1 = cv2.threshold(g, int(thr), 255, cv2.THRESH_BINARY)
                                variants.append((f"thr_{thr}", thr, bin1))
                            except Exception:
                                pass
                            try:
                                clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
                                g_clahe = clahe.apply(g)
                                _t, bin2 = cv2.threshold(g_clahe, int(thr), 255, cv2.THRESH_BINARY)
                                variants.append((f"clahe_thr_{thr}", thr, bin2))
                            except Exception:
                                pass

                        return variants

                    gray_np = np.array(crop_img)
                    variants = _plate_preprocess_variants(gray_np)
                    plate_pipeline_attempts = []

                    best_score = -1.0
                    best_meta = None
                    best_candidate_results = []
                    chosen_plate_attempt_num = None

                    for idx, (pname, pthr, img_np) in enumerate(variants):
                        try:
                            candidate_results = _ocr_read_plate(img_np)
                            score, matches = _score_plate_results(candidate_results)
                            plate_raw_groups = []
                            for (bbox, text, prob) in candidate_results:
                                height = int(
                                    ((bbox[3][1] - bbox[0][1]) + (bbox[2][1] - bbox[1][1])) / 2
                                )
                                try:
                                    pval = float(prob)
                                except Exception:
                                    pval = 0.0
                                plate_raw_groups.append((str(text), height, pval))
                            plate_pipeline_attempts.append(
                                {
                                    "attempt": idx + 1,
                                    "pipeline": pname,
                                    "threshold": pthr,
                                    "score": round(score, 2),
                                    "matches": matches,
                                    "raw_groups": plate_raw_groups,
                                    "selected": False,
                                }
                            )

                            if idx == 0:
                                best_results = candidate_results
                                selected_threshold = pthr if pthr is not None else selected_threshold
                                selected_plate_pipeline = pname

                            if score > best_score:
                                best_score = score
                                best_meta = (pname, pthr, img_np)
                                best_candidate_results = candidate_results

                            if matches >= 2:
                                successful_results = candidate_results
                                selected_plate_pipeline = pname
                                selected_threshold = pthr if pthr is not None else selected_threshold
                                chosen_plate_attempt_num = idx + 1
                                break
                        except Exception as retry_err:
                            print(f"[extract_text_from_pdf] Plate pipeline OCR error pipeline={pname} thr={pthr}: {retry_err}")

                    if successful_results:
                        results = successful_results
                    else:
                        results = best_candidate_results or best_results
                        extracted_data["_plate_retry_failed"] = True
                        if best_meta is not None and plate_pipeline_attempts:
                            try:
                                chosen_plate_attempt_num = (
                                    plate_pipeline_attempts.index(
                                        next(
                                            a
                                            for a in plate_pipeline_attempts
                                            if a["pipeline"] == best_meta[0]
                                            and a.get("threshold") == best_meta[1]
                                        )
                                    )
                                    + 1
                                )
                            except (StopIteration, ValueError):
                                chosen_plate_attempt_num = None
                        print(
                            f"[extract_text_from_pdf] Plate: ни одна попытка не дала 2 валидных части "
                            f"({len(plate_pipeline_attempts)} пайплайнов). "
                            f"Взят лучший по score, pipeline={selected_plate_pipeline}, thr={selected_threshold}."
                        )

                    if chosen_plate_attempt_num is None and plate_pipeline_attempts:
                        chosen_plate_attempt_num = len(plate_pipeline_attempts)
                    for att in plate_pipeline_attempts:
                        att["selected"] = att.get("attempt") == chosen_plate_attempt_num

                    try:
                        if best_meta is not None:
                            pname, pthr, best_np = best_meta
                            selected_plate_pipeline = selected_plate_pipeline or pname
                            crop_img = Image.fromarray(best_np)
                    except Exception:
                        crop_img = crop_img.point(lambda p: 255 if p > selected_threshold else 0)

                    extracted_data["_plate_pipeline"] = selected_plate_pipeline
                    extracted_data["_plate_threshold"] = selected_threshold
                    extracted_data["_field_logs"]["Гос_номер ()"] = {
                        "attempts": plate_pipeline_attempts,
                        "cleaning_header": _plate_cleaning_header(threshold_options, min_height_plate),
                        "winning_pipeline": selected_plate_pipeline,
                        "winning_threshold": selected_threshold,
                        "chosen_attempt": chosen_plate_attempt_num,
                        "retry_failed": bool(extracted_data.get("_plate_retry_failed")),
                    }
                    plate_ocr_precomputed = True
            
            write_preview = save_preview_files and _is_preview_photo_field(field_name)
            img_filename = get_safe_filename(pdf_path, field_name) if write_preview else None
            field_img_filename = img_filename
            img_path = os.path.join(save_dir, img_filename) if img_filename else None

            if reader is None:
                if img_path:
                    print(f"[extract_text_from_pdf] OCR reader is not initialized. Skipping OCR for {img_path}")
                results = []
            else:
                try:
                    if not plate_ocr_precomputed:
                        results = _easyocr_read_pil(crop_img)
                except Exception as e:
                    print(f"[extract_text_from_pdf] OCR error for field '{field_name}': {e}")
                    import traceback
                    traceback.print_exc()
                    results = []

            if write_preview and img_path:
                try:
                    _save_preview_jpeg(crop_img, img_path)
                    if preview_workspace_root:
                        fk = _preview_field_key(field_name)
                        if fk is not None:
                            rel = _preview_relpath(img_path, preview_workspace_root)
                            extracted_data.setdefault("_preview_by_field", {}).setdefault(
                                str(fk), []
                            ).append(rel)
                except Exception as e:
                    print(f"[extract_text_from_pdf] Не удалось сохранить JPEG предпросмотра {img_path}: {e}")
            text_parts = []
            raw_items = []
            
            min_height = CURRENT_MIN_HEIGHT_CONFIG.get(field_name, 0)
            
            for (bbox, text, prob) in results:
                height = int(((bbox[3][1] - bbox[0][1]) + (bbox[2][1] - bbox[1][1])) / 2)
                
                raw_items.append((text, height))
                extracted_data["ocr_debug"].setdefault(field_name, []).append(
                    {
                        "text": str(text),
                        "height": height,
                        "min_height": int(min_height),
                        "passed": bool(height >= min_height),
                    }
                )

                if height >= min_height:
                    text_parts.append(text)
                else:
                    pass
                    # print(f"[extract_text_from_pdf] Filtered out text '{text}' in '{field_name}' due to height {height} < {min_height}")
            
            if any(x in field_name for x in ["ФИО Водит.", "Марка", "Гос_номер"]):
                extracted_data[field_name] = raw_items
            else:
                extracted_data[field_name] = " ".join(text_parts)
                extracted_data[field_name] = " ".join(text_parts).strip()

            mode_lbl = extract_mode_label or ("кэш страницы" if used_page_cache else "полный OCR")
            if "Марка" in field_name:
                filtered_brand = DataCleaner.get_cleaned_big_3_list(raw_items)
                _record_field_attempt(
                    extracted_data,
                    field_name,
                    {
                        "ocr_mode": mode_lbl,
                        "cleaning": _brand_cleaning_desc(),
                        "raw_groups": list(raw_items),
                        "filtered_groups": filtered_brand,
                        "result": " / ".join(filtered_brand) if filtered_brand else "",
                    },
                )
            if "ФИО Водит." in field_name:
                fio_fmt_log, fio_cl_log = DataCleaner.clean_fio_raw(raw_items, {})
                _record_field_attempt(
                    extracted_data,
                    field_name,
                    {
                        "ocr_mode": mode_lbl,
                        "threshold": fio_threshold,
                        "cleaning": _fio_cleaning_desc(fio_threshold),
                        "raw_groups": list(raw_items),
                        "result": fio_cl_log,
                        "formatted": fio_fmt_log,
                    },
                )

            if timing_collector and timing_block and field_img_filename:
                field_t1 = perf_counter()
                timing_collector.add_step(
                    timing_block,
                    step_idx,
                    f"OCR поля «{field_name}»"
                    + (f" + JPEG ({field_img_filename})" if field_img_filename else " (RAM)"),
                    field_t1 - field_t0,
                )
                step_idx += 1
    except Exception as e:
        print(f"Error processing {pdf_path}: {e}")
    finally:
        if doc is not None:
            doc.close()
        if timing_collector and timing_block:
            timing_collector.set_block_total(timing_block, perf_counter() - timing_total_start)

    return extracted_data
def extract_data_from_xlsx(xlsx_path):
    extracted_data = {}
    try:
        print(f"[extract_data_from_xlsx] Loading xlsx {xlsx_path}")
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        sheet = wb.active
        
        def get_val(cell_ref):
            val = sheet[cell_ref].value
            return str(val).strip() if val is not None else ""

        date_val = get_val("K75")
        if not date_val:
            date_val = get_val("K76")
        extracted_data["Дата (1)"] = date_val

        extracted_data["Марка_XLSX"] = get_val("G89")

        plate_1 = get_val("B89")
        plate_2 = get_val("B90")
        plate_1_clean = DataCleaner.clean_plate_text(plate_1) if plate_1 else ""
        plate_2_clean = DataCleaner.clean_plate_text(plate_2) if plate_2 else ""
        extracted_data["Гос.номер_XLSX"] = f"{plate_1_clean} / {plate_2_clean}"

        extracted_data["ФИО Водит. (4)"] = get_val("M80")

        extracted_data["Кол.тон (7)"] = get_val("U43")

        wb.close()
    except Exception as e:
        print(f"[extract_data_from_xlsx] Error processing XLSX {xlsx_path}: {e}")
    
    source_map = {
        1: "K75/K76",
        2: "G89",
        3: "B89/B90",
        4: "M80",
        7: "U43"
    }
    
    return extracted_data, source_map

def process_zip_file(zip_file, dollar_rate, selected_date, tn_ved_code, bnd_code, nds_percent, save_photos=False):
    """
    Обрабатывает ZIP. Возвращает (results, preview_workspace).

    preview_workspace — временная папка в системном temp (не media/temp_ocr):
    только кропы для страницы предпросмотра; удаляется после скачивания Excel.
    """
    global CURRENT_MIN_HEIGHT_CONFIG

    legacy_media_temp = os.path.join(settings.MEDIA_ROOT, "temp_ocr")
    if os.path.isdir(legacy_media_temp):
        shutil.rmtree(legacy_media_temp, ignore_errors=True)
        print(f"[process_zip_file] Удалена устаревшая папка: {legacy_media_temp}")

    with _ocr_temp_workspace() as base_temp_dir:
        return _process_zip_file_in_workspace(
            base_temp_dir,
            zip_file,
            dollar_rate,
            selected_date,
            tn_ved_code,
            bnd_code,
            nds_percent,
            save_photos,
        )


def _process_zip_file_in_workspace(
    base_temp_dir,
    zip_file,
    dollar_rate,
    selected_date,
    tn_ved_code,
    bnd_code,
    nds_percent,
    save_photos,
):
    upload_dir = os.path.join(base_temp_dir, "upload")
    extract_dir = os.path.join(base_temp_dir, "extracted")
    imgs_root_dir = os.path.join(settings.MEDIA_ROOT, "imgs")
    preview_imgs_dir = os.path.join(base_temp_dir, "preview_imgs")
    selected_maps = get_maps_by_zip_name(getattr(zip_file, "name", ""))
    CURRENT_MIN_HEIGHT_CONFIG = selected_maps["min_height"]
    timing_collector = _TimingCollector(enabled=True)
    process_total_start = perf_counter()
    type_1_map = selected_maps["type_1"]
    type_2_map = selected_maps["type_2"]
    type_3_map = selected_maps["type_3"]
    type_2_page_2_map = selected_maps["type_2_page_2"]

    if save_photos:
        os.makedirs(imgs_root_dir, exist_ok=True)
    os.makedirs(upload_dir, exist_ok=True)
    os.makedirs(extract_dir, exist_ok=True)
    os.makedirs(preview_imgs_dir, exist_ok=True)

    zip_path = os.path.join(upload_dir, "upload.zip")
    with open(zip_path, 'wb+') as destination:
        total_written = 0
        for chunk in zip_file.chunks():
            try:
                destination.write(chunk)
                total_written += len(chunk)
            except Exception as e:
                print(f"[process_zip_file] Error writing chunk to {zip_path}: {e}")
        print(f"[process_zip_file] Wrote zip to {zip_path}, bytes={total_written}")

    print(f"Extracting ZIP to: {extract_dir}")
    with zipfile.ZipFile(zip_path, 'r') as zip_ref:
        zip_ref.extractall(extract_dir)

    type_1_files = []
    type_2_files = []
    type_3_files = []

    print("Scanning extracted files...")
    for root, dirs, files in os.walk(extract_dir):
        for file in files:
            full_path = os.path.join(root, file)
            
            if re.match(r'^[\d\.]+(\s*(cmp|смп|смр|cmr))?\s*\.(pdf|xlsx)$', file, re.IGNORECASE):
                type_1_files.append(full_path)
                print(f" [SCAN] Found Type 1 (Main): {file}")
            elif file.lower().endswith('.pdf'):
                if re.match(r'^(эсф|электронный\s*(-)?\s*счет\s*(-)?\s*фактура)', file.lower()):
                    type_2_files.append(full_path)
                    print(f" [SCAN] Found Type 2 (ESF): {file}")
                elif re.match(r'^(снт|сопроводительная\s*накладная\s*(на)?\s*товары)', file.lower()):
                    type_3_files.append(full_path)
                    print(f" [SCAN] Found Type 3 (SNT): {file}")
                else:
                    # print(f" [SCAN] Ignored PDF: {file}")
                    pass
            else:
                # print(f" [SCAN] Ignored file: {file}")
                pass


    # print(f"[process_zip_file] Found {len(type_1_files)} type_1, {len(type_2_files)} type_2, {len(type_3_files)} type_3 files")

    if not type_1_files:
        print(f"[process_zip_file] No type_1 files found in {extract_dir}")
        raise Exception("В архиве не найдены файлы основных документов (например, '1.pdf' или '1.xlsx'). Проверьте структуру архива.")

    final_results = []
    
    used_type_2 = set()
    used_type_3 = set()
    
    driver_debug_info = [] # Store debug string for each driver

    for obj_idx, t1_path in enumerate(type_1_files):
        print(f"Processing Type 1 file: {t1_path} (Basename: {os.path.basename(t1_path)})")
        
        preview_obj_dir = os.path.join(preview_imgs_dir, f"obj_{obj_idx}")
        os.makedirs(preview_obj_dir, exist_ok=True)
        t1_preview_dir = os.path.join(preview_obj_dir, "type1")
        if os.path.exists(t1_preview_dir):
            shutil.rmtree(t1_preview_dir)
        os.makedirs(t1_preview_dir, exist_ok=True)
        
        is_xlsx = t1_path.lower().endswith('.xlsx')
        t1_data = {}
        
        if is_xlsx:
            t1_data, source_map = extract_data_from_xlsx(t1_path)
            if isinstance(source_map, dict):
                source_map.pop(1, None)
            fio_raw_data = t1_data.get("ФИО Водит. (4)", [])
            
            fio_str = str(fio_raw_data).strip()
            fio_formatted = fio_str
            fio_clean = fio_str
            surname_full = fio_clean
            surname_clean = sanitize_surname(surname_full.split()[0].strip()) if surname_full else "Unknown"
            if not surname_clean:
                surname_clean = "Unknown"
        else:
            source_map = {}
            # Сначала пробуем более "жёсткую" очистку (выше порог), чтобы убрать влияние штампа/печати.
            fio_thresholds = [170, 173, 175, 178, 180, 185, 190, 168, 165, 162, 160, 155, 150, 130]
            page_cache_holder = {}
            
            best_fio_formatted = ""
            best_fio_clean = ""
            best_surname_clean = "Unknown"
            best_surname_full = "Unknown"
            
            t2_match = False
            t3_match = False
            all_field_logs = {}
            fio_meta_attempts = []
            chosen_fio_attempt = None
            
            for attempt, thresh in enumerate(fio_thresholds):
                if attempt == 0:
                    print(f"[PROCESS] Попытка {attempt+1} для {os.path.basename(t1_path)}, порог {thresh} (полный OCR)")
                    t1_data = extract_text_from_pdf(
                        t1_path,
                        type_1_map,
                        t1_preview_dir,
                        apply_deskew=True,
                        fio_threshold=thresh,
                        page_cache_holder=page_cache_holder,
                        timing_collector=timing_collector,
                        extract_mode_label="полный OCR (все поля: марка, номер, ФИО, …)",
                        preview_workspace_root=base_temp_dir,
                    )
                    _merge_field_logs(all_field_logs, t1_data.get("_field_logs"))
                else:
                    print(f"[PROCESS] Попытка {attempt+1} для {os.path.basename(t1_path)}, порог {thresh} (только ФИО, RAM-кэш страницы)")
                    partial_map = {"ФИО Водит. (4)": type_1_map["ФИО Водит. (4)"]}
                    anchor_off = t1_data.get("_anchor_offset")
                    if not isinstance(anchor_off, (list, tuple)) or len(anchor_off) != 2:
                        anchor_off = (0, 0)
                    anchor_off = (int(anchor_off[0]), int(anchor_off[1]))
                    partial_data = extract_text_from_pdf(
                        t1_path,
                        partial_map,
                        t1_preview_dir,
                        apply_deskew=True,
                        fio_threshold=thresh,
                        page_cache_image=page_cache_holder.get("image"),
                        anchor_offset_cached=anchor_off,
                        timing_collector=timing_collector,
                        extract_mode_label="только ФИО (RAM-кэш, марка/номер не перечитываются)",
                        preview_workspace_root=base_temp_dir,
                    )
                    _merge_field_logs(all_field_logs, partial_data.get("_field_logs"))
                    if "ФИО Водит. (4)" in partial_data:
                        t1_data["ФИО Водит. (4)"] = partial_data["ФИО Водит. (4)"]
                    od = partial_data.get("ocr_debug")
                    if isinstance(od, dict) and od.get("ФИО Водит. (4)"):
                        t1_data.setdefault("ocr_debug", {})["ФИО Водит. (4)"] = od["ФИО Водит. (4)"]
                
                fio_raw_data = t1_data.get("ФИО Водит. (4)", [])
                fio_formatted, fio_clean = DataCleaner.clean_fio_raw(fio_raw_data, {})

                fio_rules_passed = DataCleaner.fio_rules_ok(fio_clean)
                fio_meta_attempts.append(
                    {
                        "attempt": attempt + 1,
                        "threshold": thresh,
                        "mode": "полный OCR" if attempt == 0 else "только ФИО",
                        "raw_groups": list(fio_raw_data) if isinstance(fio_raw_data, list) else [],
                        "fio_clean": fio_clean,
                        "rules_passed": fio_rules_passed,
                        "t2_match": False,
                        "t3_match": False,
                    }
                )
                surname_full = fio_clean
                surname_clean = sanitize_surname(surname_full.split()[0].strip()) if surname_full else "Unknown"
                if not surname_clean:
                    surname_clean = "Unknown"
                
                # Если ФИО "похоже на мусор" (штамп/лишние знаки/точки) — не принимаем и пробуем следующую попытку.
                if not fio_rules_passed:
                    bad_chars = "".join(sorted({ch for ch in DataCleaner.FIO_DISALLOWED_CHARS if ch in str(fio_clean)}))
                    print(f"[PROCESS] ФИО правила не прошли (attempt={attempt+1}, порог={thresh}, dots={str(fio_clean).count('.')}, bad_chars='{bad_chars}'). Повтор OCR.")
                    fio_meta_attempts[-1]["rules_fail_reason"] = f"точек={str(fio_clean).count('.')}, bad_chars={bad_chars!r}"
                    continue

                # Зафиксируем best только после того, как ФИО прошло правила.
                if attempt == 0 or not best_fio_clean:
                    best_fio_formatted = fio_formatted
                    best_fio_clean = fio_clean
                    best_surname_clean = surname_clean
                    best_surname_full = surname_full
                
                # Если первая попытка была "плохая", но текущая проходит правила — обновим best на будущее fallback.
                if attempt != 0 and best_fio_clean and not DataCleaner.fio_rules_ok(best_fio_clean):
                    best_fio_formatted = fio_formatted
                    best_fio_clean = fio_clean
                    best_surname_clean = surname_clean
                    best_surname_full = surname_full
                
                t2_match = False
                t3_match = False
                matched_filename_surname = None
                
                if surname_clean and surname_clean != "Unknown":
                    svariants = normalize_surname(surname_clean)
                    for t2p in type_2_files:
                        if t2p in used_type_2: continue
                        fname_surname = extract_surname_from_filename(t2p)
                        if fname_surname:
                            for v in svariants:
                                ratio = difflib.SequenceMatcher(None, v.lower(), fname_surname.lower()).ratio()
                                if ratio >= 0.60:
                                    t2_match = True
                                    matched_filename_surname = fname_surname
                                    break
                            if t2_match: break
                        
                    for t3p in type_3_files:
                        if t3p in used_type_3: continue
                        fname_surname = extract_surname_from_filename(t3p)
                        if fname_surname:
                            for v in svariants:
                                ratio = difflib.SequenceMatcher(None, v.lower(), fname_surname.lower()).ratio()
                                if ratio >= 0.60:
                                    t3_match = True
                                    if not matched_filename_surname:
                                        matched_filename_surname = fname_surname
                                    break
                            if t3_match: break
                
                fio_meta_attempts[-1]["t2_match"] = t2_match
                fio_meta_attempts[-1]["t3_match"] = t3_match

                if t2_match or t3_match:
                     if matched_filename_surname:
                         fio_formatted = matched_filename_surname
                         fio_clean = matched_filename_surname
                         surname_clean = matched_filename_surname
                         surname_full = matched_filename_surname
                     best_fio_formatted = fio_formatted
                     best_fio_clean = fio_clean
                     best_surname_clean = surname_clean
                     best_surname_full = surname_full
                     chosen_fio_attempt = attempt + 1
                     print(f" [MATCH FOUND] на попытке {attempt+1} с фамилией из файла: {matched_filename_surname}")
                     break
            
            if chosen_fio_attempt is None and fio_meta_attempts:
                for meta in reversed(fio_meta_attempts):
                    if meta.get("rules_passed"):
                        chosen_fio_attempt = meta.get("attempt")
                        break
                if chosen_fio_attempt is None:
                    chosen_fio_attempt = fio_meta_attempts[-1].get("attempt")

            fio_formatted = best_fio_formatted
            fio_clean = best_fio_clean
            surname_clean = best_surname_clean
            surname_full = best_surname_full
            
            # --- Fallback to ESF/SNT filenames ---
            if not t2_match and not t3_match:
                if len(type_2_files) == len(type_3_files):
                     for f_t2 in type_2_files:
                         if f_t2 in used_type_2: continue
                         for f_t3 in type_3_files:
                             if f_t3 in used_type_3: continue
                             
                             n2 = extract_name_from_filename(f_t2)
                             n3 = extract_name_from_filename(f_t3)
                             
                             def get_overlap(t1, t2):
                                 w1 = set(re.findall(r'\w+', str(t1).lower()))
                                 w2 = set(re.findall(r'\w+', str(t2).lower()))
                                 if not w1 or not w2: return 0
                                 return len(w1.intersection(w2)) / max(len(w1), len(w2))
                             
                             # Для fallback лучше сравнивать уже "очищенную" фамилию (без мусорных символов).
                             base_for_fallback = best_surname_clean if best_surname_clean else best_surname_full
                             r2 = difflib.SequenceMatcher(None, str(base_for_fallback).lower(), n2.lower()).ratio()
                             r3 = difflib.SequenceMatcher(None, str(base_for_fallback).lower(), n3.lower()).ratio()
                             
                             if r2 > 0.6 or r3 > 0.6 or get_overlap(base_for_fallback, n2) >= 0.5:
                                 matched_surname = extract_surname_from_filename(f_t2) or extract_surname_from_filename(f_t3)
                                 fio_formatted = matched_surname if matched_surname else n2.title()
                                 fio_clean = fio_formatted
                                 surname_full = fio_formatted
                                 surname_clean = sanitize_surname(surname_full.split()[0].strip()) if surname_full else "Unknown"
                                 if not surname_clean:
                                     surname_clean = "Unknown"
                                 print(f" [FALLBACK] Взято имя из файла ЭСФ/СНТ: {fio_formatted}")
                                 t2_match = True
                                 break
                         if t2_match: break

        context = {
            'surname': surname_clean,
            'zip_filename': zip_file.name,
            'type_2_files': type_2_files,
            'type_3_files': type_3_files,
            'filename': os.path.basename(t1_path)
        }

        found_t2 = False
        found_t3 = False

        raw_date = t1_data.get("Дата (1)")
        date_clean = DataCleaner.clean_1(raw_date, context)
        if not date_clean:
            date_clean = "Unknown_Date"
        
        date_folder = date_clean.replace("/", "-").replace("\\", "-")

        person_img_dir = os.path.join(imgs_root_dir, date_folder, surname_clean)
        
        preview_obj_dir = os.path.join(preview_imgs_dir, f"obj_{len(final_results)}")
        os.makedirs(preview_obj_dir, exist_ok=True)
        preview_image_paths = []
        field_images = {}
        if not is_xlsx:
            _merge_preview_paths_from_extract(field_images, t1_data)
        _scan_dir_field_images(
            t1_preview_dir, base_temp_dir, field_images, preview_image_paths
        )

        if save_photos:
            os.makedirs(person_img_dir, exist_ok=True)
            if os.path.exists(t1_preview_dir):
                for img_file in os.listdir(t1_preview_dir):
                    shutil.copy2(os.path.join(t1_preview_dir, img_file), os.path.join(person_img_dir, img_file))
        
        plate_val = ""
        car_val = ""
        big_3_cleaned_list = []
        plate_groups = []
        plate_format_warning = False
        plate_retry_failed = bool(t1_data.get("_plate_retry_failed"))
        raw_groups = {}
        ocr_debug_merged = {}

        if is_xlsx:
            plate_val = t1_data.get("Гос.номер_XLSX", "")
            car_val = t1_data.get("Марка_XLSX", "")
            big_3_cleaned_list = [car_val, plate_val]
            if plate_val:
                parts = [p.strip() for p in str(plate_val).split("/") if p.strip()]
                plate_groups = parts
                # Требуем 2 части (до/после "/") и соответствие шаблону для каждой части
                if len(parts) != 2 or not all(DataCleaner.is_plate_token_like(p) for p in parts):
                    plate_format_warning = True
        else:
            # --- Process BRAND ---
            brand_raw = t1_data.get("Марка", [])
            # For brand, we just take all text found in the box
            raw_brand_str = " ".join([t.strip() for t, h in brand_raw if t.strip()])
            car_val = clean_brand_name(raw_brand_str)
            
            # --- Process PLATE ---
            plate_raw = t1_data.get("Гос_номер ()", [])
            plate_cleaned_list = DataCleaner.get_cleaned_big_3_list(plate_raw)
            plate_groups = list(plate_cleaned_list)
            
            # Debug: Capture filtered items with heights for Plate
            filtered_debug = []
            for text, h in plate_raw:
                if 35 < h < 55:
                     filtered_debug.append(f"{text} (H: {h})")
            filtered_debug_str = " | ".join(filtered_debug)
            
            if plate_cleaned_list:
                valid_plates = [t for t in plate_cleaned_list if DataCleaner.is_plate_token_like(t)]
                if len(valid_plates) >= 2:
                    p1 = DataCleaner.clean_plate_text(valid_plates[0])
                    p2 = DataCleaner.clean_plate_text(valid_plates[-1])
                    plate_val = f"{p1} / {p2}"
                elif len(valid_plates) == 1:
                    plate_val = DataCleaner.clean_plate_text(valid_plates[0])
                else:
                    if len(plate_cleaned_list) >= 2:
                        p1 = DataCleaner.clean_plate_text(plate_cleaned_list[0])
                        p2 = DataCleaner.clean_plate_text(plate_cleaned_list[-1])
                        plate_val = f"{p1} / {p2}"
                    else:
                        plate_val = DataCleaner.clean_plate_text(plate_cleaned_list[0])

            # Validate extracted plate against expected pattern (each side)
            parts = [p.strip() for p in str(plate_val).split("/") if p.strip()]
            if len(parts) != 2 or not all(DataCleaner.is_plate_token_like(p) for p in parts):
                plate_format_warning = True
            if plate_retry_failed:
                plate_format_warning = True

            # Raw groups from OCR before filtering (для предпросмотра)
            try:
                ocr_debug_merged.update(t1_data.get("ocr_debug", {}) if isinstance(t1_data.get("ocr_debug"), dict) else {})
            except Exception:
                pass

            def _texts_for_field(fname: str):
                items = ocr_debug_merged.get(fname) or []
                return [str(i.get("text", "")).strip() for i in items if str(i.get("text", "")).strip()]

            raw_groups[2] = _texts_for_field("Марка")
            raw_groups[3] = _texts_for_field("Гос_номер ()")
            raw_groups[4] = _texts_for_field("ФИО Водит. (4)")
            raw_groups[7] = _texts_for_field("Кол.тон (7)")

            ocr_details_text = format_driver_ocr_details(
                os.path.basename(t1_path),
                all_field_logs,
                fio_meta_attempts=fio_meta_attempts,
                final_brand=car_val,
                final_plate=plate_val,
                final_fio=fio_formatted,
                chosen_fio_attempt=chosen_fio_attempt,
            )
            print(ocr_details_text)
            driver_debug_info.append(ocr_details_text)
        
        if selected_date:
            user_date_str = selected_date.strftime('%d.%m.%Y') if hasattr(selected_date, 'strftime') else str(selected_date)
        else:
            user_date_str = ""
        
        print(
            f"[process_zip_file] field_images (type1): ключи {list(field_images.keys())}, "
            f"файлов: {sum(len(v) for v in field_images.values())}"
        )

        row_data = {
            1: user_date_str,
            2: car_val,
            3: plate_val,
            4: fio_formatted,
            5: tn_ved_code,
            6: bnd_code,
            7: DataCleaner.clean_7(t1_data.get("Кол.тон (7)"), context),
            8: None, 9: None, 10: dollar_rate,
            11: None,
            12: None, 13: None,
            14: DataCleaner.clean_14(None, context),
            15: None, 16: None,
            15: None, 16: None,
            17: "",
            18: filtered_debug_str if not is_xlsx else "",
            "plate_groups": plate_groups,
            "plate_format_warning": bool(plate_format_warning),
            "plate_retry_failed": bool(plate_retry_failed),
            "raw_groups": raw_groups,
            "ocr_debug": ocr_debug_merged,
            'preview_images': preview_image_paths,
            'field_images': field_images,
            'sources': source_map,
            'errors': []
        }

        if not is_xlsx:
            brand_raw = t1_data.get("Марка", [])
            plate_raw = t1_data.get("Гос_номер ()", [])
            
            raw_details_parts = []
            if brand_raw:
                 raw_details_parts.append(f"Brand: {' '.join([t for t, h in brand_raw])}")
            if plate_raw:
                 plate_debug = " | ".join([f"{t} (H: {h})" for t, h in plate_raw])
                 raw_details_parts.append(f"Plate: {plate_debug}")
            
            row_data[17] = " | ".join(raw_details_parts)

        if surname_clean and surname_clean != "Unknown":
            surname_variants = normalize_surname(surname_clean)
            print(f" [MATCH DEBUG] Generated variants for '{surname_clean}': {surname_variants}")
            # print(f"[process_zip_file] Searching for ESF/SNT files with surname variants: {surname_variants}")
            
            for t2_path in type_2_files:
                if t2_path in used_type_2:
                    continue
                fname_surname = extract_surname_from_filename(t2_path)
                match_found = False
                
                if fname_surname:
                    for variant in surname_variants:
                        ratio = difflib.SequenceMatcher(None, variant.lower(), fname_surname.lower()).ratio()
                        if ratio >= 0.60:
                            match_found = True
                            print(f" [MATCH] 60%+ similarity match found Type 2: {t2_path} (variant: {variant}, fname_surname: {fname_surname}, ratio: {ratio:.2f})")
                            row_data[4] = fname_surname
                            break
                    if match_found:
                        pass
                
                if match_found:
                    print(f" Match confirmed for Type 2: {t2_path}")
                    t2_preview_dir = os.path.join(preview_obj_dir, "type2")
                    os.makedirs(t2_preview_dir, exist_ok=True)
                    
                    t2_data = extract_text_from_pdf(
                        t2_path,
                        type_2_map,
                        t2_preview_dir,
                        timing_collector=timing_collector,
                        preview_workspace_root=base_temp_dir,
                    )
                    _merge_preview_paths_from_extract(field_images, t2_data)
                    try:
                        if isinstance(t2_data.get("ocr_debug"), dict):
                            ocr_debug_merged.update(t2_data["ocr_debug"])
                    except Exception:
                        pass
                    
                    price_raw = t2_data.get("Цена (8)")
                    price_val_str = DataCleaner.clean_8(price_raw, context)
                    
                    check_price = safe_decimal(price_val_str, "Check Price")
                    
                    if check_price == Decimal("7") or check_price <= Decimal("1"):
                        print(f" [Price Check] Price is {check_price}, checking 2nd page of ESF...")
                        t2_data_p2 = extract_text_from_pdf(
                            t2_path,
                            type_2_page_2_map,
                            t2_preview_dir,
                            page_num=1,
                            timing_collector=timing_collector,
                            preview_workspace_root=base_temp_dir,
                        )
                        _merge_preview_paths_from_extract(field_images, t2_data_p2)
                        price_alt_raw = t2_data_p2.get("Цена (8) Alt")
                        if price_alt_raw:
                            print(f" [Price Check] Found price on 2nd page: {price_alt_raw}")
                            t2_data["Цена (8)"] = price_alt_raw
                        else:
                            print(" [Price Check] No price found on 2nd page.")

                    row_data[8] = DataCleaner.clean_8(t2_data.get("Цена (8)"), context)
                    row_data[16] = DataCleaner.clean_16(t2_data.get("№ счет факт (Инвойс) (16)"), context)
                    
                    if os.path.exists(t2_preview_dir):
                        t2_files = [f for f in os.listdir(t2_preview_dir) if os.path.isfile(os.path.join(t2_preview_dir, f))]
                        print(f"[process_zip_file] t2_preview_dir: {len(t2_files)} файлов")
                        for img_file in t2_files:
                            img_path = os.path.join(t2_preview_dir, img_file)
                            if os.path.isfile(img_path):
                                rel_path = _preview_relpath(img_path, base_temp_dir)
                                preview_image_paths.append(rel_path)
                                
                                img_file_lower = img_file.lower()
                                
                                # Важно: не используем общие ключевые слова типа "schet/faktura",
                                # т.к. они часто встречаются в базовом имени PDF и могут попасть
                                # в filenames для ПАРА (цена/инвойс) и перепутать классификацию.
                                # Опираться будем в основном на маркеры поля: "_16" / "(16)" / "16.png".
                                has_16 = (
                                    "_16" in img_file_lower
                                    or "(16)" in img_file_lower
                                    or "16.png" in img_file_lower
                                    or "_16_" in img_file_lower
                                )
                                
                                has_8 = (
                                    "_8" in img_file_lower
                                    or "(8)" in img_file_lower
                                    or "8.png" in img_file_lower
                                    or "_8_" in img_file_lower
                                    or ("cena" in img_file_lower or "цена" in img_file_lower or
                                        "price" in img_file_lower or "tsena" in img_file_lower)
                                )
                                
                                if has_8:
                                    field_images.setdefault("8", []).append(rel_path)
                                elif has_16:
                                    pass
                                else:
                                    field_images.setdefault("8", []).append(rel_path)
                    
                    if save_photos:
                        os.makedirs(person_img_dir, exist_ok=True)
                        if os.path.exists(t2_preview_dir):
                            for img_file in os.listdir(t2_preview_dir):
                                src_path = os.path.join(t2_preview_dir, img_file)
                                dst_path = os.path.join(person_img_dir, f"type2_{img_file}")
                                shutil.copy2(src_path, dst_path)
                    
                    found_t2 = True
                    used_type_2.add(t2_path)
                    break 
                if found_t2: break
            
            if not found_t2:
                row_data['errors'].append("Не найден файл ЭСФ (Счет-фактура) для этого водителя.")
                print(f"[process_zip_file] Warning: no Type2 (ЭСФ) match for surname {surname_clean} in object {len(final_results)}")
            
            for t3_path in type_3_files:
                if t3_path in used_type_3:
                    continue
                fname_surname = extract_surname_from_filename(t3_path)
                match_found = False
                
                if fname_surname:
                    for variant in surname_variants:
                        ratio = difflib.SequenceMatcher(None, variant.lower(), fname_surname.lower()).ratio()
                        if ratio >= 0.60:
                            match_found = True
                            print(f" [MATCH] 60%+ similarity match found Type 3: {t3_path} (variant: {variant}, fname_surname: {fname_surname}, ratio: {ratio:.2f})")
                            row_data[4] = fname_surname
                            break
                    if match_found:
                        pass

                if match_found:
                    print(f" Match confirmed for Type 3: {t3_path}")
                    t3_preview_dir = os.path.join(preview_obj_dir, "type3")
                    os.makedirs(t3_preview_dir, exist_ok=True)
                    
                    t3_data = extract_text_from_pdf(
                        t3_path, type_3_map, t3_preview_dir, timing_collector=timing_collector
                    )
                    try:
                        if isinstance(t3_data.get("ocr_debug"), dict):
                            ocr_debug_merged.update(t3_data["ocr_debug"])
                    except Exception:
                        pass
                    row_data[15] = DataCleaner.clean_15(t3_data.get("№ сопров.накл. KZ (15)"), context)
                    
                    date_sopr = DataCleaner.clean_1(t3_data.get("Дата сопр.накл (13)"), context)
                    row_data[13] = date_sopr
                        
                    if os.path.exists(t3_preview_dir):
                        t3_files_list = []
                        for img_file in os.listdir(t3_preview_dir):
                            img_path = os.path.join(t3_preview_dir, img_file)
                            if os.path.isfile(img_path) and img_file.lower().endswith(('.png', '.jpg', '.jpeg')):
                                rel_path = _preview_relpath(img_path, base_temp_dir)
                                preview_image_paths.append(rel_path)
                                t3_files_list.append((img_file, rel_path))
                        print(f"[process_zip_file] t3_preview_dir: {len(t3_files_list)} png")
                        
                        for img_file, rel_path in t3_files_list:
                            img_file_lower = img_file.lower()
                            
                            has_kz_15 = ("kz" in img_file_lower or "_15" in img_file_lower or "(15)" in img_file or 
                                            "15.png" in img_file_lower or "_15_" in img_file_lower or
                                            "soprovozhdenie" in img_file_lower or "soprovozhd" in img_file_lower)
                            
                            has_13 = ("_13" in img_file_lower or "(13)" in img_file or 
                                        "13.png" in img_file_lower or "_13_" in img_file_lower)
                            has_date_keyword = ("date" in img_file_lower or "data" in img_file_lower or "дата" in img_file_lower or
                                                "datar" in img_file_lower) 
                            has_sopr_nakl = ("sopr" in img_file_lower or "сопров" in img_file_lower or "сопр" in img_file_lower or
                                                "nakl" in img_file_lower or "накл" in img_file_lower)
                            
                            if has_kz_15 and not has_13:
                                if 15 not in field_images:
                                    field_images[15] = []
                                field_images[15].append(rel_path)
                            elif (has_13 or (has_date_keyword and has_sopr_nakl)) and not has_kz_15:
                                if 13 not in field_images:
                                    field_images[13] = []
                                field_images[13].append(rel_path)
                            elif 13 not in field_images and 15 not in field_images:
                                file_index = [i for i, (f, _) in enumerate(t3_files_list) if f == img_file][0]
                                if file_index == 0:
                                    if 15 not in field_images:
                                        field_images[15] = []
                                    field_images[15].append(rel_path)
                                else:
                                    if 13 not in field_images:
                                        field_images[13] = []
                                    field_images[13].append(rel_path)
                            elif 15 in field_images and 13 not in field_images:
                                if 13 not in field_images:
                                    field_images[13] = []
                                field_images[13].append(rel_path)
                            elif 13 in field_images and 15 not in field_images:
                                if 15 not in field_images:
                                    field_images[15] = []
                                field_images[15].append(rel_path)
                    
                    if save_photos:
                        os.makedirs(person_img_dir, exist_ok=True)
                        if os.path.exists(t3_preview_dir):
                            for img_file in os.listdir(t3_preview_dir):
                                src_path = os.path.join(t3_preview_dir, img_file)
                                dst_path = os.path.join(person_img_dir, f"type3_{img_file}")
                                shutil.copy2(src_path, dst_path)
                    
                    found_t3 = True
                    used_type_3.add(t3_path)
                    break 
                if found_t3: break
            
            if not found_t3:
                row_data['errors'].append("Не найден файл СНТ (Накладная) для этого водителя.")
                print(f"[process_zip_file] Warning: no Type3 (СНТ) match for surname {surname_clean} in object {len(final_results)}")
        else:
            print(f"[process_zip_file] Surname not found or empty ('{surname_clean}'), skipping ESF/SNT matching.")

        try:
            kol_ton = safe_decimal(row_data[7], "Кол.тон (7)")
            kol_ton = kol_ton / Decimal("1000")
            row_data[7] = kol_ton
            
            cena = safe_decimal(row_data[8], "Цена (8)")
            row_data[8] = cena
            
            sum_dollar = (kol_ton * cena).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            row_data[9] = sum_dollar
            
            sum_som = (sum_dollar * dollar_rate).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            row_data[11] = sum_som
            
            nds_percent_value = nds_percent if isinstance(nds_percent, Decimal) else Decimal(str(nds_percent))
            nds_sum = (sum_som * nds_percent_value / Decimal("100")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            row_data[12] = nds_sum
            
            if found_t3 and not row_data.get(15):
                row_data['errors'].append("Не удалось найти '№ сопров.накл. KZ'. Проверьте файл СНТ.")
            if found_t2 and not row_data.get(16):
                row_data['errors'].append("Не удалось найти '№ счет факт'. Проверьте файл ЭСФ.")
            if found_t3 and not row_data.get(13):
                row_data['errors'].append("Не удалось найти 'Дата сопр.накл'. Проверьте файл СНТ.")
            
        except Exception as e:
            print(f"[process_zip_file] Calculation error for object {len(final_results)}: {e}")

        final_results.append(row_data)

    unused_t2 = len(type_2_files) - len(used_type_2)
    unused_t3 = len(type_3_files) - len(used_type_3)

    # --- Force Match Logic (1-1-1 Rule) ---
    incomplete_rows = []
    for i, res in enumerate(final_results):
        has_esf = res.get(16) is not None
        has_snt = res.get(15) is not None
        if not has_esf or not has_snt:
             incomplete_rows.append(i)

    if len(incomplete_rows) == 1 and unused_t2 == 1 and unused_t3 == 1:
        row_idx = incomplete_rows[0]
        row_data = final_results[row_idx]
        
        print(f"[Force Match] Triggered! 1 incomplete row (Index {row_idx}), 1 unused ESF, 1 unused SNT.")

        # Находим неиспользованные файлы
        leftover_t2 = list(set(type_2_files) - used_type_2)[0]
        leftover_t3 = list(set(type_3_files) - used_type_3)[0]
        
        force_surname = extract_surname_from_filename(leftover_t2) or extract_surname_from_filename(leftover_t3)
        if force_surname:
            row_data[4] = force_surname
            print(f"[Force Match] Updated driver name to matched surname: {force_surname}")
        
        print(f"[Force Match] Force linking ESF: {os.path.basename(leftover_t2)}")
        print(f"[Force Match] Force linking SNT: {os.path.basename(leftover_t3)}")
        
        # Reconstruct context
        t1_path_for_row = type_1_files[row_idx]
        fio_fmt = row_data.get(4, "")
        surname_for_ctx = sanitize_surname(fio_fmt.split()[0].strip()) if fio_fmt else "Unknown"
        if not surname_for_ctx:
            surname_for_ctx = "Unknown"
        
        context = {
            'surname': surname_for_ctx,
            'zip_filename': zip_file.name,
            'type_2_files': type_2_files,
            'type_3_files': type_3_files,
            'filename': os.path.basename(t1_path_for_row)
        }

        # Directories
        target_preview_dir = os.path.join(preview_imgs_dir, f"obj_{row_idx}")
        t2_preview_dir = os.path.join(target_preview_dir, "type2")
        t3_preview_dir = os.path.join(target_preview_dir, "type3")
        os.makedirs(t2_preview_dir, exist_ok=True)
        os.makedirs(t3_preview_dir, exist_ok=True)

        # 1. Extract ESF (Type 2)
        t2_data = extract_text_from_pdf(
            leftover_t2,
            type_2_map,
            t2_preview_dir,
            timing_collector=timing_collector,
            preview_workspace_root=base_temp_dir,
        )
        _merge_preview_paths_from_extract(row_data["field_images"], t2_data)
        
        price_raw = t2_data.get("Цена (8)")
        price_val_str = DataCleaner.clean_8(price_raw, context)
        check_price = safe_decimal(price_val_str, "Check Price")
        
        if check_price == Decimal("7") or check_price <= Decimal("1"):
            print(f" [Force Match] checking 2nd page of ESF... (Price={check_price})")
            t2_data_p2 = extract_text_from_pdf(
                leftover_t2,
                type_2_page_2_map,
                t2_preview_dir,
                page_num=1,
                timing_collector=timing_collector,
                preview_workspace_root=base_temp_dir,
            )
            _merge_preview_paths_from_extract(row_data["field_images"], t2_data_p2)
            price_alt_raw = t2_data_p2.get("Цена (8) Alt")
            if price_alt_raw:
                t2_data["Цена (8)"] = price_alt_raw

        row_data[8] = DataCleaner.clean_8(t2_data.get("Цена (8)"), context)
        row_data[16] = DataCleaner.clean_16(t2_data.get("№ счет факт (Инвойс) (16)"), context)
        used_type_2.add(leftover_t2)

        # 2. Extract SNT (Type 3)
        t3_data = extract_text_from_pdf(
            leftover_t3, type_3_map, t3_preview_dir, timing_collector=timing_collector
        )
        row_data[15] = DataCleaner.clean_15(t3_data.get("№ сопров.накл. KZ (15)"), context)
        date_sopr = DataCleaner.clean_1(t3_data.get("Дата сопр.накл (13)"), context)
        row_data[13] = date_sopr
        used_type_3.add(leftover_t3)
        
        # 3. Update Images in row_data
        def scan_and_add_images(scan_dir, r_data):
            if os.path.exists(scan_dir):
                for img_file in os.listdir(scan_dir):
                    img_path = os.path.join(scan_dir, img_file)
                    if os.path.isfile(img_path):
                        rel_path = _preview_relpath(img_path, base_temp_dir)
                        r_data['preview_images'].append(rel_path)
                        
                        img_lower = img_file.lower()
                        # Здесь тоже не используем общие ключевые слова (schet/invoice),
                        # чтобы не перепутать цену и инвойс из-за base имени PDF.
                        has_16 = ("_16" in img_lower or "(16)" in img_lower or "16.png" in img_lower or "_16_" in img_lower)
                        has_8 = ("_8" in img_lower or "(8)" in img_lower or "8.png" in img_lower or "_8_" in img_lower or
                                  ("cena" in img_lower or "price" in img_lower or "tsena" in img_lower))
                        if has_16:
                             if 16 not in r_data['field_images']: r_data['field_images'][16] = []
                             r_data['field_images'][16].append(rel_path)
                        elif has_8:
                             if 8 not in r_data['field_images']: r_data['field_images'][8] = []
                             r_data['field_images'][8].append(rel_path)
                        else:
                             if 8 not in r_data['field_images']: r_data['field_images'][8] = []
                             r_data['field_images'][8].append(rel_path)
                        
                        has_kz_15 = ("kz" in img_lower or "_15" in img_lower or "(15)" in img_lower)
                        has_13 = ("_13" in img_lower or "(13)" in img_lower or "date" in img_lower)
                        if has_kz_15:
                             if 15 not in r_data['field_images']: r_data['field_images'][15] = []
                             r_data['field_images'][15].append(rel_path)
                        elif has_13:
                             if 13 not in r_data['field_images']: r_data['field_images'][13] = []
                             r_data['field_images'][13].append(rel_path)

        scan_and_add_images(t2_preview_dir, row_data)
        scan_and_add_images(t3_preview_dir, row_data)

        # 4. Save photos if needed
        if save_photos:
            sname = context['surname']
            r_date = row_data.get(1, "Unknown_Date")
            r_date_folder = str(r_date).replace("/", "-").replace("\\", "-")
            person_img_dir = os.path.join(imgs_root_dir, r_date_folder, sname)
            os.makedirs(person_img_dir, exist_ok=True)
            for d in [t2_preview_dir, t3_preview_dir]:
                if os.path.exists(d):
                    for img_file in os.listdir(d):
                         shutil.copy2(os.path.join(d, img_file), os.path.join(person_img_dir, img_file))

        # 5. Clear missing file errors
        new_errors = []
        for err in row_data['errors']:
             if "Не найден файл" in err or "Не удалось найти" in err:
                 continue
             new_errors.append(err)
        row_data['errors'] = new_errors
        
        # 6. Recalculate Totals
        try:
             # Ensure kol_ton is Decimal
             kt = row_data.get(7)
             if not isinstance(kt, Decimal):
                 kt = safe_decimal(kt, "Кол.тон (7)")
                 # Note: in loop, row_data[7] was result of / 1000 if successful.
                 # If it failed before, it might be string. 
                 # If it was successful, it is Decimal (tons).
                 # We assume if it was 20000kg -> 20t. 
                 # If it is > 500, likely it is still in kg?
                 if kt > 500: 
                     kt = kt / Decimal("1000")
                 row_data[7] = kt
             
             price = safe_decimal(row_data.get(8), "Цена (8)")
             row_data[8] = price
             
             sum_dollar = (kt * price).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
             row_data[9] = sum_dollar
             
             sum_som = (sum_dollar * dollar_rate).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
             row_data[11] = sum_som
             
             n_val = nds_percent if isinstance(nds_percent, Decimal) else Decimal(str(nds_percent))
             nds_sum = (sum_som * n_val / Decimal("100")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
             row_data[12] = nds_sum
             
        except Exception as e:
            print(f"[Force Match] Re-calculation error: {e}")
            row_data['errors'].append(f"Ошибка пересчета после Force Match: {e}")

        unused_t2 = 0
        unused_t3 = 0

    if unused_t2 > 0 or unused_t3 > 0:
        print(f"[process_zip_file] Unused files: unused_t2={unused_t2}, unused_t3={unused_t3}")
        
        # Build detailed diagnostics
        debug_msg = "\n--- ДЕТАЛИЗАЦИЯ ОБРАБОТКИ ---\n"
        for i, res in enumerate(final_results):
            surname = res.get(4, "Unknown")
            has_esf = "OK" if res.get(16) else "MISSING"
            has_snt = "OK" if res.get(15) or res.get(13) else "MISSING" # 15=KZ num, 13=Date
            debug_msg += f"#{i+1}: {surname} | ЭСФ: {has_esf} | СНТ: {has_snt}\n"
        
        unused_t2_files = sorted(list(set(type_2_files) - used_type_2))
        unused_t3_files = sorted(list(set(type_3_files) - used_type_3))
        
        if unused_t2_files:
            debug_msg += "\nНеиспользованные файлы ЭСФ:\n" + "\n".join([os.path.basename(f) for f in unused_t2_files])
        if unused_t3_files:
            debug_msg += "\nНеиспользованные файлы СНТ:\n" + "\n".join([os.path.basename(f) for f in unused_t3_files])

        print(debug_msg) # Ensure it goes to console logs

        error_msg = "Обнаружено несоответствие количества файлов:\n"
        if unused_t2 > 0:
            error_msg += f"- Лишних файлов ЭСФ (Счет-фактура): {unused_t2} шт.\n"
        if unused_t3 > 0:
            error_msg += f"- Лишних файлов СНТ (Накладная): {unused_t3} шт.\n"
        error_msg += "Убедитесь, что для каждого ЭСФ/СНТ есть соответствующий основной документ (PDF/XLSX).\n"
        error_msg += "\nСм. детали в логах ниже (какие файлы остались, какие водители обработаны):\n" + debug_msg
        
        if driver_debug_info:
             error_msg += "\n\n--- ДЕТАЛИЗАЦИЯ ПО ВОДИТЕЛЯМ (РАСПОЗНАВАНИЕ) ---\n" + "\n".join(driver_debug_info)

        raise Exception(error_msg)

    _cleanup_ocr_processing_artifacts(base_temp_dir)
    timing_collector.print_all(overall_seconds=perf_counter() - process_total_start)
    return final_results, base_temp_dir


def generate_excel(data, existing_excel_file=None, nds_percent=2):
    has_numbering_column = False
    next_row_number = 1
    
    if existing_excel_file:
        wb = openpyxl.load_workbook(existing_excel_file)
        ws = wb.active
        
        first_cell_value = ws.cell(row=1, column=1).value
        if first_cell_value:
            first_cell_str = str(first_cell_value).lower()
            if "дата" not in first_cell_str:
                has_numbering_column = True
                print(f"[generate_excel] Detected numbering column. First header: '{first_cell_value}'")
                
                max_number = 0
                for row_idx in range(2, ws.max_row + 1):
                    cell_value = ws.cell(row=row_idx, column=1).value
                    if cell_value is not None:
                        try:
                            num = int(cell_value)
                            if num > max_number:
                                max_number = num
                        except (ValueError, TypeError):
                            pass
                next_row_number = max_number + 1
                print(f"[generate_excel] Will continue numbering from {next_row_number}")
            else:
                print(f"[generate_excel] First column contains 'дата', no numbering column detected")
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "OCR Results"
        headers = [
            "Дата", "Марка АТС", "Гос.номер АТС", "ФИО Водит.", "Код ТН ВЭД",
            "БНД", "Кол.тон", "Цена", "Сумма в $", "Курс", "Сумма в сомах",
            "НДС ЕАЭС", "Дата сопр.накл", "Номер СМР", "№ сопров.накл. KZ", "№ счет факт"
        ]
        ws.append(headers)

    bold_font = Font(bold=True)
    header_row = 1
    max_col = ws.max_column
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.font = bold_font

    fill_light_green = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    fill_dark_green = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")

    for row in data:
        if has_numbering_column:
            row_values = [
                next_row_number,
                row.get(1), row.get(2), row.get(3), row.get(4), row.get(5),
                row.get(6), row.get(7), row.get(8), None,
                row.get(10), None,
                None,
                row.get(13), row.get(14), row.get(15), row.get(16)
            ]
            next_row_number += 1
        else:
            row_values = [
                row.get(1), row.get(2), row.get(3), row.get(4), row.get(5),
                row.get(6), row.get(7), row.get(8), None,
                row.get(10), None,
                None,
                row.get(13), row.get(14), row.get(15), row.get(16)
            ]
        ws.append(row_values)
        
        current_row = ws.max_row
        
        col_offset = 1 if has_numbering_column else 0
        
        from openpyxl.utils import get_column_letter
        
        col_kol_ton = get_column_letter(7 + col_offset)
        col_price = get_column_letter(8 + col_offset)
        col_sum_dollar = get_column_letter(9 + col_offset)
        col_rate = get_column_letter(10 + col_offset)
        col_sum_som = get_column_letter(11 + col_offset)
        col_nds = get_column_letter(12 + col_offset)
        
        ws[f'{col_sum_dollar}{current_row}'] = f'={col_kol_ton}{current_row}*{col_price}{current_row}'
        
        ws[f'{col_sum_som}{current_row}'] = f'={col_sum_dollar}{current_row}*{col_rate}{current_row}'
        
        ws[f'{col_nds}{current_row}'] = f'={col_sum_som}{current_row}*{nds_percent}%'
        
        for col_idx in range(1, 19 + col_offset):
            cell = ws.cell(row=current_row, column=col_idx)
            
            val = cell.value
            if val:
                if has_numbering_column and col_idx == 1:
                    continue
                
                field_idx = col_idx - col_offset
                
                if field_idx in [1, 13] and isinstance(val, str):
                    try:
                        dt = datetime.strptime(val, "%d.%m.%Y").date()
                        cell.value = dt
                        cell.number_format = 'DD.MM.YYYY'
                    except ValueError:
                        pass
                    
                if field_idx in [5, 14] and isinstance(val, str) and val.isdigit():
                    try:
                        cell.value = int(val)
                        cell.number_format = '0'
                    except:
                        pass
                        
        ws.cell(row=current_row, column=1 + col_offset).fill = fill_light_green
        ws.cell(row=current_row, column=10 + col_offset).fill = fill_light_green
        
        ws.cell(row=current_row, column=7 + col_offset).fill = fill_dark_green

    col_offset = 1 if has_numbering_column else 0
    numeric_cols = [7 + col_offset, 8 + col_offset, 9 + col_offset, 10 + col_offset, 11 + col_offset, 12 + col_offset]
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for col_idx in numeric_cols:
            cell = row[col_idx - 1]
            cell.number_format = '#,##0.00'

    existing_tables = list(ws.tables.values())
    for existing_table in existing_tables:
        del ws.tables[existing_table.name]
    
    ws.auto_filter = None

    return wb